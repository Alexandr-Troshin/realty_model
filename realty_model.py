import datetime

import openpyxl as ox
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.styles.borders import Side
from openpyxl.styles import PatternFill, Font, Border, Alignment
import time
from sklearn import tree
import pandas as pd
from bs4 import BeautifulSoup
import requests
import numpy as np
import re
import datetime as dt
from tqdm.auto import tqdm, trange
from copy import copy
from natasha import AddrExtractor, MorphVocab
import traceback
import pymorphy2
import pymorphy2_dicts_ru

from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

import logging
import os
# для работы progress_apply
tqdm.pandas(desc="In Progress")

# -
def is_lot_finished(driver, lot_url):
    lot_status = 0
    final_price = np.nan
    try:
        driver.get(lot_url)
        try:
            wait = WebDriverWait(driver, 3)
            wait.until(EC.presence_of_element_located((By.XPATH, #visibility_of_element_located
             "//div[@class='notice-pub-title-and-subtitle']")))
#            driver.implicitly_wait(1)
            if driver.find_element(By.XPATH, "//div[@class='notice-status with-background amsterdam']"):
                if driver.find_element(By.XPATH, "//div[@class='notice-status with-background amsterdam']").text.strip() == 'Завершено':
                    try:
                        wait = WebDriverWait(driver, 1)
                        wait.until(EC.presence_of_element_located((By.XPATH,
                         "//div[@class='lot-collapsed_header']")))
#                        driver.implicitly_wait(1)
                        try:
                            not_succeed_el = driver.find_elements(By.XPATH, "//div[@class='protocol-lot-status error']")
                        except:
                            pass
                        if not_succeed_el:
                            if not_succeed_el.text.strip() == 'Не состоялся':
                                lot_status = -1
                    #---> перенос второго if

                    # except (Exception,):
                    #     pass
                    #
                    # try:
                    #     wait = WebDriverWait(driver, 10)
                    #     waintil(EC.visibility_of_element_located((By.XPATH,
                    #      "//div[@clt.uass='lot-collapsed_header']")))
 #                       driver.implicitly_wait(1)
                        elif driver.find_element(By.XPATH, "//div[@class='protocol-lot-status amsterdam']"):
                            plsa_element = driver.find_element(By.XPATH, "//div[@class='protocol-lot-status amsterdam']")
                            if plsa_element.text.strip() == 'Состоялся':
                                lot_status = 1
                                #wait = WebDriverWait(driver, 10)
                                try:
                                    wait.until(EC.presence_of_element_located((By.XPATH,
                                    '//div[@class="lot-protocols"]/div/app-inline-icon/app-icon-chevron-right/*[name()="svg"]'))).click()
                                except:
                                    pass
                                wait.until(EC.presence_of_element_located((By.XPATH,
                                 '//span[contains(@class,"button__label") and contains(text(), "итог") '
                                 'and contains(text(), "аукцион")]'))).click()
                                # wait.until(EC.visibility_of_element_located((By.XPATH,
                                #  '//button[contains(.,"Информация")]'))).click()
                                final_price = int(re.sub(' ', '', wait.until(EC.presence_of_element_located((By.XPATH,
                                 '//div[@class="content__result__bet__value"]'))).text.split(',')[0]))
                    except:
                        pass

        except:
            pass

    except (Exception,):
        pass

    return (driver, lot_status, final_price)


def go_to_procedure_detail(driver, tab_name):
    """Функция переключения по вкладкам внутри сраницы объекта на investmoscow.ru
    :param tab_name: надпись на вкладке
    :return driver после открытия нужной вкладки"""
    element = driver.find_element(By.XPATH, "//div[text()='"+tab_name+"']")
    body = driver.find_element(By.CSS_SELECTOR,'body')

    try:
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//div[text()='"+tab_name+"']")))
        # body.send_keys(Keys.ARROW_DOWN)
        # body.send_keys(Keys.ARROW_DOWN)
        # body.send_keys(Keys.ARROW_DOWN)
        # body.send_keys(Keys.ARROW_DOWN)
        # body.send_keys(Keys.ARROW_DOWN)
        #        time.sleep(1)
        element.click()
    except (Exception,):
        try:
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//div[text()='"+tab_name+"']")))
            body.send_keys(Keys.ARROW_DOWN)
            body.send_keys(Keys.ARROW_DOWN)
            body.send_keys(Keys.ARROW_DOWN)
            #            time.sleep(1)
            element.click()
        except:
            try:
                body.send_keys(Keys.ARROW_UP)
                body.send_keys(Keys.ARROW_UP)
                body.send_keys(Keys.ARROW_UP)
                body.send_keys(Keys.ARROW_UP)
                body.send_keys(Keys.ARROW_UP)
                body.send_keys(Keys.ARROW_UP)
                #                time.sleep(1)
                element.click()
            except:
                try:
                    WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, "//div[text()='"+tab_name+"']")))
                    #                    time.sleep(1)
                    element.click()
                except:
                    print('Dont click')
                    logging.info('Dont click')
    return driver

def parse_lot(driver, url):
    """ Функция получает на вход URL лота с сайта investmoscow.ru, парсит данный лот и возвращает pd.Series
        с характеристиками лота.
        :param url: url лота
        :return: возвращает Pandas Series с данными по лоту.
        """

    # создаем словарь объекта
    obj_dict = {LOT_FIELDS[i]: np.nan for i in range(len(LOT_FIELDS))}
    driver.get(url)
    wait = WebDriverWait(driver, 5)
    try:
        wait.until(
            EC.visibility_of_element_located((By.XPATH,
                                          '//div[@class="tender__content"]')))
    except:
        driver.refresh()
        wait.until(
            EC.visibility_of_element_located((By.XPATH,
                                              '//div[@class="tender__content"]')))
    full_info = BeautifulSoup(driver.find_element(By.XPATH, '//div[@class="tender__content"]').get_attribute('innerHTML'),
                         "html.parser")
    obj_dict['investmoscow_url'] = url

    # выбираем данные по площади и местоположению
    try:
        obj_dict['obj_square'] = float(
            str.replace(full_info.find('span', class_="uid-text-accent uid-mr-16").text, ',', '.').split(' ')[0])
    except Exception as exc:
        print('Ошибка определения площади объекта ', url, exc)
        logging.info(f"Лот {url} - Ошибка определения площади объекта")

    labels_info = full_info.find_all('div', class_="tender-card__header-label")
    for label in labels_info:
        if label.text.strip().startswith('#'):
            try:
                obj_dict['lot_tag'] = label.text.strip()
            except Exception as exc:
                obj_dict['lot_tag'] = np.nan
                print('Ошибка определения номера лота объекта ', url, exc)
                logging.info(f"Лот {url} - Ошибка определения номера лота объекта")
        elif label.text.strip().startswith('Тип объекта:'):
            try:
                obj_dict['object_type'] = label.text.split(':')[-1].strip()
            except Exception as exc:
                obj_dict['object_type'] = np.nan
                print('Ошибка определения типа объекта ', url, exc)
                logging.info(f"Лот {url} - Ошибка определения типа объекта")
        elif label.text.strip().startswith('Вид торгов:'):
            try:
                obj_dict['auct_type'] = label.text.split(':')[-1].strip()
            except Exception as exc:
                obj_dict['auct_type'] = np.nan
                print('Ошибка определения вида торгов ', url, exc)
                logging.info(f"Лот {url} - Ошибка определения вида торгов объекта")

    try:
        obj_dict['addr_string'] = full_info.find('div', class_="tender-card__description").div.next_sibling.text
    except Exception as exc:
        print('Ошибка определения адреса объекта ', url, exc)
        logging.info(f"Лот {url} - Ошибка определения адреса объекта")

    obj_desc = full_info.find('div', class_="subject-table")
    # данные из блока "Сведения об объекте"
    # забираем только интересующие данные
    t1 = obj_desc.find_all('div', class_="subject-table-row")
    for block in t1:
        label_text = block.find('div', class_="subject-table-label").text.strip()
        if label_text == 'Этаж:':
            try:
                obj_dict['addr_floor'] = int(block.find('div', class_="subject-table-text").text
                                             .strip('№').strip('N').strip())
            except Exception as exc:
                if block.find('div', class_="subject-table-text").text != 'Не указано':
                    print(f"Лот {url} - Ошибка определения этажа объекта")
                    logging.info(f"Лот {url} - Ошибка определения этажа объекта")

        elif label_text == 'Этажность дома:':
            try:
                obj_dict['total_floors'] = int(block.find('div', class_="subject-table-text").text)
            except Exception as exc:
                if block.find('div', class_="subject-table-text").text != 'Не указано':
                    print(f"Лот {url} - Ошибка определения этажности объекта")
                    logging.info(f"Лот {url} - Ошибка определения этажности объекта")

        elif label_text == 'Тип объекта:':
            if not obj_dict['object_type']:
                obj_dict['object_type'] = block.find('div', class_="subject-table-text").text

        elif label_text == 'Кадастровый номер:':
            obj_dict['cadastr_num'] = block.find('div', class_="subject-table-text").text

        elif label_text == 'Номер квартиры:':
            try:
                obj_dict['flat_num'] = int(block.find('div', class_="subject-table-text").text
                                           .strip('№').strip('N').strip())
            except Exception as exc:
                if block.find('div', class_="subject-table-text").text != 'Не указано':
                    print(f"Лот {url} - Ошибка определения номера квартиры объекта")
                    logging.info(f"Лот {url} - Ошибка определения номера квартиры объекта")

        elif label_text == 'Количество комнат:':
            try:
                obj_dict['qty_rooms'] = int(block.find('div', class_="subject-table-text").text)
            except Exception as exc:
                if block.find('div', class_="subject-table-text").text != 'Не указано':
                    print(f"Лот {url} - Ошибка определения количества комнат объекта")
                    logging.info(f"Лот {url} - Ошибка определения количества комнат объекта")
    # переключаемся на "Сведения о процедуре"

    driver = go_to_procedure_detail(driver, 'Сведения о процедуре')


    wait.until(
        EC.visibility_of_element_located((By.XPATH,
                                          '//div[@class="tender__content"]')))
    full_info = BeautifulSoup(
        driver.find_element(By.XPATH, '//div[@class="tender__content"]').get_attribute('innerHTML'),
        "html.parser")
    obj_desc = full_info.find('div', class_="subject-table")
    t1 = obj_desc.find_all('div', class_="subject-table-row")

    for block in t1:
        label_text = block.find('div', class_="subject-table-label").text.strip()

        if label_text == 'Начальная цена за объект:':
            try:
                obj_dict['start_price'] = int(re.sub(r"\D", "",
                                            block.find('div', class_="subject-table-text").text.split(
                                                         ',')[0]))
                obj_dict['start_price_m2'] = round(obj_dict['start_price'] / obj_dict['obj_square'], 2)
            except Exception as exc:
                print(f"Лот {url} - Ошибка определения стартовой цены объекта")
                logging.info(f"Лот {url} - Ошибка определения стартовой цены объекта")

        elif label_text == 'Размер задатка:':
            try:
                obj_dict['deposit'] = int(re.sub(r"\D", "",
                                            block.find('div', class_="subject-table-text").text.split(
                                                         ',')[0]))
            except Exception as exc:
                print(f"Лот {url} - Ошибка определения задатка объекта")
                logging.info(f"Лот {url} - Ошибка определения задатка объекта")

        elif label_text == 'Шаг аукциона:':
            try:
                obj_dict['auct_step'] = int(re.sub(r"\D", "",
                                            block.find('div', class_="subject-table-text").text.split(
                                                         ',')[0]))
            except Exception as exc:
                print(f"Лот {url} - Ошибка определения шага аукциона объекта")
                logging.info(f"Лот {url} - Ошибка определения шага аукциона объекта")

        elif label_text == 'Форма проведения:':
            obj_dict['auct_form'] = block.find('div', class_="subject-table-text").text

        elif label_text == 'Дата начала приёма заявок:':
            obj_dict['start_applications_date'] = pd.to_datetime(
                block.find('div', class_="subject-table-text").text.split()[0],
                infer_datetime_format=True, format='%d.%m.%Y', errors='coerce')

        elif label_text == 'Дата окончания приёма заявок:':
            obj_dict['finish_application_date'] = pd.to_datetime(
                block.find('div', class_="subject-table-text").text.split()[0],
                infer_datetime_format=True, format='%d.%m.%Y', errors='coerce')

        elif label_text == 'Отбор участников:':
            obj_dict['participant_selection_date'] = pd.to_datetime(
                block.find('div', class_="subject-table-text").text.split()[0],
                infer_datetime_format=True, format='%d.%m.%Y', errors='coerce')

        elif label_text == 'Проведение торгов:':
            obj_dict['bidding_date'] = pd.to_datetime(
                block.find('div', class_="subject-table-text").text.split()[0],
                infer_datetime_format=True, format='%d.%m.%Y', errors='coerce')

        elif label_text == 'Подведение итогов:':
            obj_dict['results_date'] = pd.to_datetime(
                block.find('div', class_="subject-table-text").text.split()[0],
                infer_datetime_format=True, format='%d.%m.%Y', errors='coerce')

        elif label_text == 'Ссылка на ЭТП:':
            try:
                obj_dict['roseltorg_url'] = block.find('a')['href']
            except Exception as exc:
                pass

        elif label_text == 'Ссылка на torgi.gov.ru:':
            try:
                obj_dict['torgi_url'] = block.find('a')['href']
            except Exception as exc:
                pass


    # переключаемся на "Дополнительная информация"
    try:
        driver = go_to_procedure_detail(driver, 'Дополнительная информация')

        wait.until(
            EC.visibility_of_element_located((By.XPATH,
                                              '//div[@class="tender__content"]')))
        full_info = BeautifulSoup(
            driver.find_element(By.XPATH, '//div[@class="extra-info"]').get_attribute('innerHTML'),
            "html.parser")
        #ищем блок "Транспортная доступность", проверяем наличие картинки метро, сохраняем название станции
        # блок с названием метро исключаем - уточнять будем по flatinfo
        # trans_desc = full_info.find('div', class_="extra-info-transport")
        # metro_logo = trans_desc.find('img')['src']
        # if re.search('metro', metro_logo):
        #     obj_dict['metro_station'] = trans_desc.div.div.div.text

        # ищем блок "Инфраструктура"
        infra_desc = full_info.find('div', class_="extra-info-block")
        t1 = infra_desc.find_all('a', class_="extra-info-block-item uid-portal-link")
        for block in t1:
            label_text = block.find('span').text.strip()
            infra_item_meaning = int(block.find('div', class_="extra-info-block-item-text").text.strip())
            if label_text == "Торговля":
                obj_dict['inf_sales'] = infra_item_meaning
            elif label_text == "Общественное питание":
                obj_dict['inf_food_service'] = infra_item_meaning
            elif label_text == "Образование":
                obj_dict['inf_education'] = infra_item_meaning
            elif label_text == "Культура и спорт":
                obj_dict['inf_cult_and_sport'] = infra_item_meaning
            elif label_text == "Бытовое обслуживание":
                obj_dict['inf_consumer_services'] = infra_item_meaning
            elif label_text == "Здравоохранение":
                obj_dict['inf_health_care'] = infra_item_meaning
            else:
                print(f"Лот {url} - неучтенная инфраструктура - ", label_text, infra_item_meaning)
                logging.info(f"Лот {url} - неучтенная инфрастурктура - "
                             + label_text + str(infra_item_meaning))
    except Exception as exc:
        print(f"Лот {url} - вкладка допинформации отсутствует")
        logging.info(f"Лот {url} - вкладка допинформации отсутствует")

    return driver, pd.Series(obj_dict)

# -
def control_lot(driver, url):
    """ Функция получает на вход URL лота с сайта investmoscow.ru.
        :param url: url лота
        :return: возвращает Pandas Series с данными по лоту.
        """

    # создаем словарь объекта
    obj_dict = dict({'status': np.nan,
                     'final_price': np.nan,
                     'final_price_m2': np.nan,
                     'delta_price': np.nan
                     })

    driver.get(url)
    wait = WebDriverWait(driver, 5)
    try:
        wait.until(
            EC.visibility_of_element_located((By.XPATH,
                                          '//div[@class="tender__content"]')))
    except:
        driver.refresh()
        try:
            wait.until(
                EC.visibility_of_element_located((By.XPATH,
                                                  '//div[@class="tender__content"]')))
        except:
            pass
    try:
        obj_dict['status'] = driver.find_element(By.XPATH, '//div[@class="tender-status__text"]').text.strip()
    except:
        obj_dict['status'] = 'failed_to_parse'

    if obj_dict['status'] == 'ПРИЗНАНЫ СОСТОЯВШИМИСЯ':
        # переключаемся на "Сведения о процедуре"
        #---> версия до 17.12.22
        # wait.until(
        #     EC.element_to_be_clickable((By.XPATH,
        #                                 "//div[@class='subject-tabs']/ul/li[2]")))
        # element = driver.find_element(By.XPATH, "//div[@class='subject-tabs']/ul/li[2]")
        # try:
        #     element.click()
        # except:
        #     pass
        # ActionChains(driver).move_to_element(
        #     WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH,
        #                                       "//div[@class='subject-tabs']/ul/li[2]")))).click().perform()
        #-------<версия до 17.12.22

        # ---> версия переключения на сведения о процедуре от 17.12.22
        go_to_procedure_detail(driver, 'Сведения о процедуре')

        wait.until(
            EC.visibility_of_element_located((By.XPATH,
                                              '//div[@class="tender__content"]')))
        full_info = BeautifulSoup(
            driver.find_element(By.XPATH, '//div[@class="tender__content"]').get_attribute('innerHTML'),
            "html.parser")
        obj_desc = full_info.find('div', class_="subject-table")
        t1 = obj_desc.find_all('div', class_="subject-table-row")
        for block in t1:
            label_text = block.find('div', class_="subject-table-label").text.strip()
            # при флаге прошедших - добавляется итоговая цена + цена за м2 + дельта

            if label_text == 'Итоговая цена:':
                obj_dict['final_price'] = int(
                    re.sub(r"\D", "", block.find('div', class_="subject-table-text").text.split(',')[0]))


        # try:
        #     obj_dict['final_price_m2'] = round(obj_dict['final_price'] / obj_dict['obj_square'], 2)
        #     obj_dict['delta_price'] = round(
        #         (obj_dict['final_price'] - obj_dict['start_price']) / obj_dict['start_price'], 4)
        # except (Exception,):
        #     obj_dict['final_price_m2'] = np.nan
        # if (obj_dict['final_price_m2'] == 0) or (pd.isnull(obj_dict['final_price_m2'])):
        #     obj_dict['final_price_m2'] = 'не состоялись(проверка)'
    return driver, pd.Series(obj_dict)


def primary_addr_normalize(line):
    """
    Функция первичной нормализации адреса. Получает на вход строку адреса и пытается ее распарсить,
    не углубляясь в детали. Срабатывает на большую часть стандартно оформленных адресов.
    Возвращает список из строки с указанием города/улицы и строки с указанием дома/корпуса/строения
    """
    adr_street = ''
    adr_house = ''
    adr_floor = np.nan
    adr_apart_num = np.nan
    for _, paramValue in enumerate(addr_extractor.find(line).fact.parts):
        p0 = addr_extractor.find(line).fact.parts
    try:
        for p in p0:
            if p.type == 'город':
                if p.value != 'Москва' and p.value != 'Москвы':
                    adr_street += str(', г. ' + p.value)
            elif p.type == 'деревня':
                adr_street += str(', д. ' + p.value)
            elif p.type == 'микрорайон':
                adr_street += str(', мкр. ' + p.value)
            elif p.type == 'бульвар':
                adr_street += str(', б-р. ' + p.value)
            elif p.type == 'дачный поселок':
                adr_street += str(', дп. ' + p.value)
            elif p.type == 'шоссе':
                adr_street += str(', ш. ' + p.value)
            elif p.type == 'улица':
                adr_street += str(', ул. ' + p.value)
            elif p.type == 'переулок':
                adr_street += str(', пер. ' + p.value)
            elif p.type == 'квартал':
                adr_street += str(', кв-л. ' + p.value)
            elif p.type == 'проезд':
                adr_street += str(', проезд. ' + p.value)
            elif p.type == 'проспект':
                adr_street += str(', пр-кт. ' + p.value)
            elif p.type == 'аллея':
                adr_street += str(', аллея. ' + p.value)
            elif p.type == 'площадь':
                adr_street += str(', пл. ' + p.value)
            elif p.type == 'набережная':
                adr_street += str(', наб. ' + p.value)
            elif p.type == 'село':
                adr_street += str(', с. ' + p.value)
            elif p.type == 'дом':
                adr_house = str(', д. ' + p.value)
            elif p.type == 'корпус':
                adr_house = adr_house + str(', к. ' + p.value)
            elif p.type == 'строение':
                adr_house = adr_house + str(', стр. ' + p.value)
            elif p.type == 'этаж':
                try:
                    adr_floor = int(p.value.strip('№').strip('N').strip())
                except Exception as exc:
                    print(line,'ошибка определения этажа')
            elif p.type == 'квартира':
                try:
                    adr_apart_num = int(p.value.strip('№').strip('N').strip())
                except Exception as exc:
                    print(line,'ошибка определения номера квартиры')
        adr_street = 'г. Москва' + adr_street
        ret = [adr_street, adr_house, adr_floor, adr_apart_num]
    except:
        ret = ['не распознан', 'не распознан', 'не распознан', 'не распознан']
    return ret


def secondary_addr_normalize(line):
    """
    Функция углубленной нормализации адреса. Получает на вход строку адреса и пытается ее распарсить,
    при этом проверяет вхождение слов из словаря conv_dictionary.xlsx и заменяет на соответствующие,
    переносит нумерацию улиц в конец строки (например, 1-я Трудовая должна быть "Трудовая 1-я ул."),
    обрабатывает названия Большой, Нижний и т.д., заменяет букву "ё" на "е".
    Так как делает больше проверок, запускается на части датафрейма, которую не удалось распарсить певично.
    Возвращает список из строки с указанием города/улицы и строки с указанием дома/корпуса/строения
    """
    # morph_vocab = MorphVocab()
    # addr_extractor = AddrExtractor(morph_vocab)
    adr_street = ''
    adr_house = ''
    adr_floor = np.nan
    adr_apart_num = np.nan
        # загружаем свой словарь обработки нестандартных адресов
    # TODO: заменить на csv
    conv_df = pd.read_excel(r'..\realty_model_files\conv_dictionary.xlsx', sheet_name='dict', engine='openpyxl')
    for con_quest in conv_df.parsed_name:
        if re.search(con_quest, line) is not None:
            print(con_quest)
            print(conv_df[conv_df.parsed_name == con_quest].correct_name)
            logging.info(str(con_quest))
            logging.info(str(conv_df[conv_df.parsed_name == con_quest].correct_name))
            line = line.replace(con_quest, conv_df[conv_df.parsed_name == con_quest].correct_name.tolist()[0])
    if re.search(r'\d+-[я] улица', line) is not None:
        st = re.search(r'\d+-[я] улица', line)
        st1 = st[0].split()
        line = re.sub(st[0], st1[1] + ' ' + st1[0], line)
    if re.search(r'\d+-[й] квартал', line) is not None:
        st = re.search(r'\d+-[й] квартал', line)
        st1 = st[0].split()
        line = re.sub(st[0], st1[1] + ' ' + st1[0], line)
    for _, paramValue in enumerate(addr_extractor.find(line).fact.parts):
        p0 = addr_extractor.find(line).fact.parts
    try:
        for p in p0:
            if p.type == 'город':
                if p.value != 'Москва' and p.value != 'Москвы':
                    adr_street += str(', г. ' + p.value)
            elif p.type == 'деревня':
                adr_street += str(', д. ' + p.value)
            elif p.type == 'микрорайон':
                adr_street += str(', мкр. ' + p.value)
            elif p.type == 'бульвар':
                adr_street += str(', б-р. ' + p.value)
            elif p.type == 'дачный поселок':
                adr_street += str(', дп. ' + p.value)
            elif p.type == 'поселок':
                adr_street += str(', п. ' + p.value)
            elif p.type == 'шоссе':
                adr_street += str(', ш. ' + p.value)
            elif p.type == 'квартал':
                adr_street += str(', кв-л. ' + p.value)
            elif p.type == 'улица':
                str_cons = p.value.split(' ')
                if str_cons[0] == '43':
                    p.value = '43-й Армии'
                if str_cons[-1] in ('Б', 'М', 'Нов', 'Стар', 'Нижн', 'Верхн', 'Ср'):
                    p.value = p.value + '.'
                if len(str_cons) > 1 and re.search(r'\d+-[я]', str_cons[0]) is not None:
                    adr = ''
                    for i in range(1, len(str_cons)):
                        adr += str_cons[i] + ' '
                    adr_street += str(', ул. ' + adr + str_cons[0])
                elif len(str_cons) > 1 and re.search('Больш', str_cons[0]) is not None:
                    adr = ''
                    for i in range(1, len(str_cons)):
                        adr += str_cons[i] + ' '
                    adr_street += str(', ул. ' + adr + 'Б.')
                elif len(str_cons) > 1 and re.search('Новая', str_cons[0]) is not None:
                    adr = ''
                    for i in range(1, len(str_cons)):
                        adr += str_cons[i] + ' '
                    adr_street += str(', ул. ' + adr + 'Нов.')
                elif len(str_cons) > 1 and re.search('Старая', str_cons[0]) is not None:
                    adr = ''
                    for i in range(1, len(str_cons)):
                        adr += str_cons[i] + ' '
                    adr_street += str(', ул. ' + adr + 'Стар.')
                elif len(str_cons) > 1 and re.search('Мал', str_cons[0]) is not None:
                    adr = ''
                    for i in range(1, len(str_cons)):
                        adr += str_cons[i] + ' '
                    adr_street += str(', ул. ' + adr + 'М.')
                elif len(str_cons) > 1 and re.search('Нижн', str_cons[0]) is not None:
                    adr = ''
                    for i in range(1, len(str_cons)):
                        adr += str_cons[i] + ' '
                    adr_street += str(', ул. ' + adr + 'Нижн.')
                elif len(str_cons) > 1 and re.search('Верхн', str_cons[0]) is not None:
                    adr = ''
                    for i in range(1, len(str_cons)):
                        adr += str_cons[i] + ' '
                    adr_street += str(', ул. ' + adr + 'Верхн.')
                else:
                    adr_street += str(', ул. ' + p.value)

            elif p.type == 'переулок':
                str_cons = p.value.split(' ')
                if str_cons[-1] in ('Б', 'М', 'Нов', 'Стар', 'Нижн', 'Верхн', 'Ср'):
                    p.value = p.value + '.'
                if len(str_cons) > 1 and re.search(r'\d+-[й]', str_cons[0]) is not None:
                    adr = ''
                    for i in range(1, len(str_cons)):
                        adr += str_cons[i] + ' '
                    adr_street += str(', пер. ' + adr + str_cons[0])
                elif len(str_cons) > 1 and re.search('Средн', str_cons[0]) is not None:
                    adr = ''
                    for i in range(1, len(str_cons)):
                        adr += str_cons[i] + ' '
                    adr_street += str(', пер. ' + adr + 'Ср.')
                elif len(str_cons) > 1 and re.search('Малый', str_cons[0]) is not None:
                    adr = ''
                    for i in range(1, len(str_cons)):
                        adr += str_cons[i] + ' '
                    adr_street += str(', пер. ' + adr + 'М.')
                elif len(str_cons) > 1 and re.search('Большой', str_cons[0]) is not None:
                    adr = ''
                    for i in range(1, len(str_cons)):
                        adr += str_cons[i] + ' '
                    adr_street += str(', пер. ' + adr + 'Б.')
                else:
                    adr_street += str(', пер. ' + p.value)

            elif p.type == 'проезд':
                str_cons = p.value.split(' ')
                if str_cons[-1] in ('Б', 'М', 'Нов', 'Стар', 'Нижн', 'Верхн', 'Ср'):
                    p.value = p.value + '.'
                if len(str_cons) > 1 and re.search(r'\d+-[й]', str_cons[0]) is not None:
                    adr = ''
                    for i in range(1, len(str_cons)):
                        adr += str_cons[i] + ' '
                    adr_street += str(', проезд. ' + adr + str_cons[0])
                elif len(str_cons) > 1 and re.search('Верхний', str_cons[0]) is not None:
                    adr = ''
                    for i in range(1, len(str_cons)):
                        adr += str_cons[i] + ' '
                    adr_street += str(', проезд. ' + adr + 'Верхн.')
                elif len(str_cons) > 1 and re.search('Больш', str_cons[0]) is not None:
                    adr = ''
                    for i in range(1, len(str_cons)):
                        adr += str_cons[i] + ' '
                    adr_street += str(', проезд. ' + adr + 'Б.')
                else:
                    adr_street += str(', проезд. ' + p.value)

            elif p.type == 'проспект':
                adr_street += str(', пр-кт. ' + p.value)
            elif p.type == 'аллея':
                adr_street += str(', аллея. ' + p.value)
            elif p.type == 'площадь':
                str_cons = p.value.split(' ')
                if str_cons[-1] in ('Б', 'М', 'Нов', 'Стар', 'Нижн', 'Верхн', 'Ср'):
                    p.value = p.value + '.'
                if len(str_cons) > 1 and re.search('Большая', str_cons[0]) is not None:
                    adr = ''
                    for i in range(1, len(str_cons)):
                        adr += str_cons[i] + ' '
                    adr_street += str(', пл. ' + adr + 'Б.')
                else:
                    adr_street += str(', пл. ' + p.value)
            elif p.type == 'набережная':
                adr_street += str(', наб. ' + p.value)
            elif p.type == 'село':
                adr_street += str(', с. ' + p.value)
            elif p.type == 'дом':
                adr_house = str(', д. ' + p.value)
            elif p.type == 'корпус':
                adr_house = adr_house + str(', к. ' + p.value)
            elif p.type == 'строение':
                adr_house = adr_house + str(', стр. ' + p.value)
            elif p.type == 'этаж':
                try:
                    adr_floor = int(p.value.strip('№').strip('N').strip())
                except Exception as exc:
                    print(line, 'ошибка определения этажа')
            elif p.type == 'квартира':
                try:
                    adr_apart_num = int(p.value.strip('№').strip('N').strip())
                except Exception as exc:
                    print(line, 'ошибка определения номера квартиры')

        adr_street = re.sub('ё', 'е', adr_street)
        adr_street = 'г. Москва' + adr_street
        ret = [adr_street, adr_house, adr_floor, adr_apart_num]
    except:
        return ['не распознан', 'не распознан', 'не распознан', 'не распознан']
    return ret

# -
def secondary_addr_normalize_for_metrodistance(line):
    """
    Функция углубленной нормализации адреса. Получает на вход строку адреса и пытается ее распарсить,
    при этом проверяет вхождение слов из словаря conv_dictionary.xlsx и заменяет на соответствующие,
    переносит нумерацию улиц в конец строки (например, 1-я Трудовая должна быть "Трудовая 1-я ул."),
    обрабатывает названия Большой, Нижний и т.д., заменяет букву "ё" на "е".
    Так как делает больше проверок, запускается на части датафрейма, которую не удалось распарсить первично.
    Возвращает список из строки с указанием города/улицы и строки с указанием дома/корпуса/строения
    """
    # morph_vocab = MorphVocab()
    # addr_extractor = AddrExtractor(morph_vocab)
    adr_street = ''
    adr_house = ''
    if re.search(r'\d+-[я] улица', line) is not None:
        st = re.search(r'\d+-[я] улица', line)
        st1 = st[0].split()
        line = re.sub(st[0], st1[1] + ' ' + st1[0], line)
    if re.search(r'\d+-[й] квартал', line) is not None:
        st = re.search(r'\d+-[й] квартал', line)
        st1 = st[0].split()
        line = re.sub(st[0], st1[1] + ' ' + st1[0], line)
    for _, paramValue in enumerate(addr_extractor.find(line).fact.parts):
        p0 = addr_extractor.find(line).fact.parts
    for p in p0:
        if p.type == 'город':
            if p.value != 'Москва' and p.value != 'Москвы':
                adr_street += str(', г. ' + p.value)
        elif p.type == 'деревня':
            adr_street += str(', д. ' + p.value)
        elif p.type == 'микрорайон':
            adr_street += str(', мкр. ' + p.value)
        elif p.type == 'бульвар':
            adr_street += str(', б-р. ' + p.value)
        elif p.type == 'дачный поселок':
            adr_street += str(', дп. ' + p.value)
        elif p.type == 'поселок':
            adr_street += str(', п. ' + p.value)
        elif p.type == 'шоссе':
            adr_street += str(', ш. ' + p.value)
        elif p.type == 'квартал':
            adr_street += str(', кв-л. ' + p.value)
        elif p.type == 'улица':
            str_cons = p.value.split(' ')
            if str_cons[0] == '43':
                p.value = '43-й Армии'
            if len(str_cons) > 1 and re.search(r'\d+-[я]', str_cons[0]) is not None:
                adr = ''
                for i in range(1, len(str_cons)):
                    adr += str_cons[i] + ' '
                adr_street += str(', ул. ' + adr + str_cons[0])
            elif len(str_cons) > 1 and re.search('Больш', str_cons[0]) is not None:
                adr = ''
                for i in range(1, len(str_cons)):
                    adr += str_cons[i] + ' '
                adr_street += str(', ул. ' + adr + 'Б.')
            elif len(str_cons) > 1 and re.search('Мал', str_cons[0]) is not None:
                adr = ''
                for i in range(1, len(str_cons)):
                    adr += str_cons[i] + ' '
                adr_street += str(', ул. ' + adr + 'М.')
            elif len(str_cons) > 1 and re.search('Нижн', str_cons[0]) is not None:
                adr = ''
                for i in range(1, len(str_cons)):
                    adr += str_cons[i] + ' '
                adr_street += str(', ул. ' + adr + 'Нижн.')
            elif len(str_cons) > 1 and re.search('Верхн', str_cons[0]) is not None:
                adr = ''
                for i in range(1, len(str_cons)):
                    adr += str_cons[i] + ' '
                adr_street += str(', ул. ' + adr + 'Верхн.')
            else:
                adr_street += str(', ул. ' + p.value)
        elif p.type == 'переулок':
            str_cons = p.value.split(' ')
            if len(str_cons) > 1 and re.search(r'\d+-[й]', str_cons[0]) is not None:
                adr = ''
                for i in range(1, len(str_cons)):
                    adr += str_cons[i] + ' '
                adr_street += str(', пер. ' + adr + str_cons[0])
            elif len(str_cons) > 1 and re.search('Средн', str_cons[0]) is not None:
                adr = ''
                for i in range(1, len(str_cons)):
                    adr += str_cons[i] + ' '
                adr_street += str(', пер. ' + adr + 'Ср.')
            else:
                adr_street += str(', пер. ' + p.value)
        elif p.type == 'проезд':
            str_cons = p.value.split(' ')
            if len(str_cons) > 1 and re.search(r'\d+-[й]', str_cons[0]) is not None:
                adr = ''
                for i in range(1, len(str_cons)):
                    adr += str_cons[i] + ' '
                adr_street += str(', проезд. ' + adr + str_cons[0])
            elif len(str_cons) > 1 and re.search('Верхний', str_cons[0]) is not None:
                adr = ''
                for i in range(1, len(str_cons)):
                    adr += str_cons[i] + ' '
                adr_street += str(', проезд. ' + adr + 'Верхн.')
            elif len(str_cons) > 1 and re.search('Больш', str_cons[0]) is not None:
                adr = ''
                for i in range(1, len(str_cons)):
                    adr += str_cons[i] + ' '
                adr_street += str(', проезд. ' + adr + 'Б.')
            else:
                adr_street += str(', проезд. ' + p.value)
        elif p.type == 'проспект':
            adr_street += str(', пр-кт. ' + p.value)
        elif p.type == 'аллея':
            adr_street += str(', аллея. ' + p.value)
        elif p.type == 'площадь':
            adr_street += str(', пл. ' + p.value)
        elif p.type == 'набережная':
            adr_street += str(', наб. ' + p.value)
        elif p.type == 'село':
            adr_street += str(', с. ' + p.value)
        elif p.type == 'дом':
            adr_house = str(', д. ' + p.value)
        elif p.type == 'корпус':
            adr_house = adr_house + str(', к. ' + p.value)
        elif p.type == 'строение':
            adr_house = adr_house + str(', стр. ' + p.value)
    adr_street = re.sub('ё', 'е', adr_street)
    adr_street = 'г. Москва' + adr_street
    ret = adr_street + adr_house
    return ret

# -
def clean_gkh_add(path: str, sheet_name: str = "Sheet1"):
    """
    Функция очистки данных в конкретном листе.
    :param path: Путь до файла Excel
    :param sheet_name: Имя листа в таблице Excel, куда буду писать данные
    :return: возвращает файл пустой файл с названиями столбцов
    """
    wb = ox.load_workbook(path)
    wb[sheet_name].delete_rows(2, 1000)
    wb.save(path)

# -
def add_unrecognized():
    """ Функция добавляет данные из файла new_buildings.xlsx в базу ЖКХ, сохраняет в csv всю базу,
        удаляет импортированные записи из файла new_buildings.xlsx.
    """
    global gkh_df


    print('Добавляются данные в базу ЖКХ')
    logging.info('Добавляются данные в базу ЖКХ')
    gkh_add = pd.read_excel(r'..\realty_model_files\new_buildings.xlsx', sheet_name='Sheet1', engine='openpyxl', keep_default_na=False)
    gkh_add.dropna(subset=['gkh_address'])
    gkh_df = gkh_df.append(gkh_add, ignore_index=True)
    gkh_df = gkh_df.drop_duplicates(subset=['gkh_address'], keep='last')
    gkh_df.reset_index(drop=True, inplace=True)
    gkh_df.to_csv(GKH_BASE_FILENAME, index=False)
    clean_gkh_add(r'..\realty_model_files\new_buildings.xlsx', sheet_name='Sheet1')


def update_gkh_base(address, res):
    """ обновление информации по адресу в базе ЖКХ.
        Не забыть скачать базу до и сохранить в файл после работы  """
    global new_gkh_df

    new_gkh_df = pd.concat([new_gkh_df, pd.DataFrame.from_records([res])] ,ignore_index=True)
    new_gkh_df.drop_duplicates(subset=['gkh_address'], keep='last', inplace=True, ignore_index=True)
    new_gkh_df.reset_index(drop=True, inplace=True)

# -
def update_new_buildings(path: str, _df, startcol: int = 1, startrow: int = 1, sheet_name: str = "Sheet1"):
    """Функция добавления данных в new_buildings.
    :param path: Путь до файла Excel
    :param _df: Датафрейм Pandas для записи
    :param startcol: Стартовая колонка в таблице листа Excel, куда буду писать данные
    :param startrow: Стартовая строка в таблице листа Excel, куда буду писать данные
    :param sheet_name: Имя листа в таблице Excel, куда буду писать данные
    :return: возвращает файл с датафреймом  (названия столбцов предустановлены),
    """
    wb = ox.load_workbook(path)
    for ir in range(0, len(_df)):
        for ic in range(0, len(_df.iloc[ir])):
            wb[sheet_name].cell(startrow + ir, startcol + ic).value = _df.iloc[ir][ic]
    wb.save(path)


def convert_addr_to_winner_type(addr_str):
    print(addr_str)
#    addr_str = re.sub("г. москва, ", "", addr_str)
    addr_els = addr_str.split(', ')
    addr_street_els = addr_els[1].strip().split(' ')

    if addr_street_els[0] == 'ул.':
        addr_street = ' '.join(addr_street_els[1:]) + ' ул.'
    elif addr_street_els[0] == 'ш.':
        addr_street = ' '.join(addr_street_els[1:]) + ' ш.'
    elif addr_street_els[0] == 'проезд.':
        addr_street = ' '.join(addr_street_els[1:]) + ' пр-д'
    elif addr_street_els[0] == 'наб.':
        addr_street = ' '.join(addr_street_els[1:]) + ' наб.'
    elif addr_street_els[0] == 'пр-кт.':
        addr_street = ' '.join(addr_street_els[1:]) + ' просп.'

    addr_build = ''
    for i in range(len(addr_els)-2):
        if addr_els[i+2].startswith('д.'):
            addr_build += addr_els[i+2].split(' ')[1]
        elif addr_els[i+2].startswith('к.'):
            addr_build += 'к' + addr_els[i+2].split(' ')[1]
        elif addr_els[i + 2].startswith('стр.'):
            addr_build += 'с' + addr_els[i + 2].split(' ')[1]

    addr_as_winner = addr_street + ', ' + addr_build
    return addr_as_winner


def metro_and_floor_data(addr_norm, url_ready, is_for_winner=False):
    """ Функция поиска данных об адресе на ресурсе flatInfo.ru (этажность и расстояние до метро)
    Функция дополнительно обновляет базу ЖКХ update_gkh_base    """
    global driver
    global gkh_df
    global new_gkh_df

    print(addr_norm)
    logging.info(str(addr_norm))
    #res = [np.nan for _ in range(5)]
    res = {GKH_FIELDS[i]: np.nan for i in range(len(GKH_FIELDS))}
    if pd.isna(url_ready):
        try:

            driver.get('https://flatinfo.ru')
            element = driver.find_element(By.XPATH, "//div[@class='search-home input-group search-home_show']/input[1]")
            adr1 = re.sub('проезд.', 'проезд', addr_norm)
            adr1 = re.sub('пр-кт.', 'проспект', adr1)
            adr1 = re.sub('пр-д', 'проспект', adr1)
            adr1 = re.sub('г. зеленоград, к.', 'г. зеленоград,', adr1)
            # ---> переворот цифровых обозначений в названии "Трудовая 1-я" -> "1-я Трудовая"
            if is_for_winner:
                if re.search(r'\d+-[я] ул.', adr1) is not None:
                    st = re.search(r'\d+-[я] улица', adr1)
                    st1 = st[0].split()
                    adr1 = re.sub(st[0], st1[1] + ' ' + st1[0], adr1)
                if re.search(r'\d+-[й] кв.', adr1) is not None:
                    st = re.search(r'\d+-[й] квартал', adr1)
                    st1 = st[0].split()
                    adr1 = re.sub(st[0], st1[1] + ' ' + st1[0], adr1)
                if re.search(r'\d+-[й] пр-д', adr1) is not None:
                    st = re.search(r'\d+-[й] проезд', adr1)
                    st1 = st[0].split()
                    adr1 = re.sub(st[0], st1[1] + ' ' + st1[0], adr1)
      # >----
            element.send_keys(adr1)
            time.sleep(3)
            butt = driver.find_element(By.XPATH, "//div[@class='search-home input-group search-home_show']/button[1]")
            ActionChains(driver).click(butt).perform()
            time.sleep(3)
            try:
                ActionChains(driver).click(butt).perform()
            except (Exception, ):
                pass

            # и сохраняем адрес перехода
            cur_url = driver.current_url
            print(cur_url)
            logging.info(str(cur_url))
            # если переход произошел на страницу с данными о доме
            if re.search('h_info', cur_url) is None:
                print('адрес не найден')
                logging.info('адрес не найден')
                print('Ввести правильный url?')
                is_cont = input()
                while is_cont not in ['Y', 'y', 'N', 'n']:
                    print('Некорректный ввод - y/n')
                    is_cont = input()

                if is_cont in ['Y', 'y']:
                    print('Введите правильный url')
                    new_url = input()
                    driver.get(new_url)
                    cur_url = new_url

            if re.search('h_info', cur_url) is not None:
                full_addr_str = driver.find_element(By.XPATH, "//h1[starts-with(text(), 'О доме')]").text
                addr_line = full_addr_str.split(' в ')[0]
                # addr_line = re.sub(' в Москве', '',full_addr_str)
                # addr_line = re.sub(' в Зеленограде', '', addr_line)
                # addr_line = re.sub(' в Зеленограде', '', addr_line)
                addr_line = re.sub(' ЗЕЛЕНОГРАД Г.', ' г. Зеленоград ', addr_line)
                parse_addr_line = secondary_addr_normalize_for_metrodistance(addr_line)
                parse_addr_line = re.sub(' г. Зеленоград, д.', ' г. Зеленоград, к.', parse_addr_line)
                print(addr_norm)
                print(parse_addr_line)
                is_equal_addr = False
                if not is_for_winner and (parse_addr_line.lower()) == addr_norm:
                    is_equal_addr = True
                elif is_for_winner and (convert_addr_to_winner_type(parse_addr_line.lower()) == addr_norm.lower()):
                    is_equal_addr = True

                print('convert ', convert_addr_to_winner_type(parse_addr_line.lower()))
                print('parse_addr', parse_addr_line.lower())
                print('start_addr', addr_norm.lower())
                print(is_equal_addr)
                if not is_equal_addr:

                    return res

                    print('адреса не совпадают')
                    print(addr_norm)
                    print(parse_addr_line)
                    logging.info('адреса не совпадают')
                    print('Ввести правильный url?')
                    is_cont = input()
                    while is_cont not in ['Y', 'y', 'N', 'n']:
                        print('Некорректный ввод - y/n')
                        is_cont = input()
                    if is_cont in ['Y', 'y']:
                        print('Введите правильный url')
                        new_url = input()
                        driver.get(new_url)
                        cur_url = new_url
                    else:
                        #driver.quit()
                        return res
                print('адреса норм')
                logging.info('адреса норм')
                if is_for_winner:
                    res['addr_winner'] = addr_norm.lower()
                    res['gkh_address'] = parse_addr_line.lower()
                else:
                    res['gkh_address'] = addr_norm
                #driver.quit()
        except Exception as exc:
            print(exc)
    else:
        cur_url = url_ready
        if is_for_winner:
            driver.get(cur_url)
            full_addr_str = driver.find_element(By.XPATH, "//h1[starts-with(text(), 'О доме')]").text
            addr_line = full_addr_str.split(' в ')[0]
            addr_line = re.sub(' ЗЕЛЕНОГРАД Г.', ' г. Зеленоград ', addr_line)
            parse_addr_line = secondary_addr_normalize_for_metrodistance(addr_line)
            parse_addr_line = re.sub(' г. Зеленоград, д.', ' г. Зеленоград, к.', parse_addr_line)
            print(addr_norm)
            print(parse_addr_line)
            if (convert_addr_to_winner_type(parse_addr_line.lower()) == addr_norm.lower()):
                res['gkh_address'] = parse_addr_line.lower()
            else:
                return res
        else:
            res['gkh_address'] = addr_norm
        # driver = start_browser_for_parse()
    try:
        try:
            r = requests.get(cur_url)
            res['building_page_url'] = cur_url
            soup = BeautifulSoup(r.text, 'lxml')
            # скачиваем страницу инфы о доме
            full_info = soup.find('div', class_='page__content')
            li_blocks = full_info.find_all('li', class_='fi-list__item fi-list-item')
            # и ищем характеристики этажности
            for li_bl in li_blocks:
                try:
                    label_text = li_bl.find('span', class_='fi-list-item__label').text.strip()
                    param_value = li_bl.find('span', class_='fi-list-item__value').text.strip()
                except:
                    continue

                if label_text == 'Этажей всего':
                    t2 = param_value.strip('\r').strip('\n').strip('\t').split('\n')
                    t3 = t2[0].strip()
                    res['gkh_total_floors'] = t3

                elif label_text == 'Год постройки':
                    try:
                        res['construction_year'] = int(param_value.strip('\n').strip())
                    except Exception as exc:
                        print(addr_norm, 'ошибка определения года постройки ', param_value)
                        logging.info(f'{addr_norm} ошибка определения года постройки {param_value}')
                elif label_text == 'Округ':
                    res['adm_area'] = param_value.strip('\n').strip()
                elif label_text == 'Район':
                    res['mun_district'] = param_value.strip('\n').strip()
                elif label_text == 'Гео-координаты':
                    try:
                        res['geo_lat'] = round(float(param_value.strip('\n')
                                                     .split('/')[0].strip()), 5)
                        res['geo_lon'] = round(float(param_value.strip('\n')
                                                     .split('/')[1].strip()), 5)
                    except Exception as exc:
                        print(addr_norm, 'ошибка определения координат ', param_value)
                        logging.info(f'{addr_norm} ошибка определения координат {param_value}')
                elif label_text == 'Перекрытия':
                    res['overlap_material'] = param_value.strip('\n').strip()
                elif label_text == 'Каркас':
                    res['skeleton'] = param_value.strip('\n').strip()
                elif label_text == 'Стены':
                    res['wall_material'] = param_value.strip('\n').strip()
                elif label_text == 'Категория':
                    res['category'] = param_value.strip('\n').strip()
                elif label_text == 'Лифтов в подъезде' or \
                        label_text == 'Пассажирских лифтов в подъезде':
                    res['passenger_elevators_qty'] = param_value.strip('\n').strip()
                elif label_text == 'Состояние':
                    res['condition'] = param_value.strip('\n').strip()
                elif label_text == 'Кадастровый номер дома':
                    res['gkh_cadastr_num'] = param_value.strip('\n').strip()
                elif label_text == 'Высота потолков':
                    try:
                        res['ceiling_height'] = int(param_value.strip('\n').strip().split(' ')[0])
                    except Exception as exc:
                        print('Ошибка определения высоты потолков - ', param_value)
                elif label_text == 'Код адреса КЛАДР':
                    res['code_KLADR'] = param_value.strip('\n').strip()
                elif label_text == 'Расселение по реновации':
                    if not re.search('не включен', param_value.lower()):
                        res['is_renov'] = 1
                        value_lst = param_value.strip('\n').strip().split(' ')
                        res['renov_period'] = value_lst[-4] + '-' + value_lst[-2]
                elif label_text == 'Проживает':
                    try:
                        res['residents_qty'] = int(param_value.strip('\n').strip().split(' ')[0])
                    except Exception as exc:
                        print(addr_norm, 'ошибка определения количества проживающих ', param_value)
                        logging.info(f'{addr_norm} ошибка определения количества проживающих {param_value}')
                else:
                    if re.search('ремонт', label_text):
                        print(addr_norm, 'упоминание капремонта  ', label_text, param_value)
                        logging.info(f'{addr_norm} упоминание капремонта {label_text} {param_value}')

            try:
                transport_data = full_info.find('ul', class_='fi-list underground')
                transp_bl = transport_data.find_all('li', class_='fi-list__item fi-list-item')
                for bl in transp_bl:
                    is_walking_distance = bl.find('svg',
                                                class_='location__how-label icon-svg').contents[1]
                    if re.search('#walking', str(is_walking_distance)):
                        res['gkh_metro_station'] = bl.find('span',
                                                           class_='fi-list-item__label').text.strip()
                        res['metro_min'] = int(bl.find('span',
                                                       class_='location__time').text
                                        .strip().split('\xa0')[0])
                        res['metro_km'] = round(float(bl.find('span',
                                                              class_='fi-list-item__value')
                                        .text.strip().split(' ')[-2])/1000, 2)
                        break
                    print()
            except Exception as exc:
                pass

        except (Exception, ):
            print('данные о доме не получены')
            logging.info('данные о доме не получены')
    except Exception as exc:
        print (exc)
    # try:
    #     driver.quit()
    # except (Exception, ):
    #     pass
    if pd.isna(res['gkh_metro_station']):
        res['gkh_metro_station'] = 'Отсутствует в пешей доступности'
        res['metro_min'] = 90
        res['metro_km'] = 20
    if pd.isna(res['is_renov']):
        res['is_renov'] = 0

    try:
        if re.search('от ', res['gkh_total_floors']) and re.search('от ', res['gkh_total_floors']):
            res['is_total_floors_variable'] = 1
    except (Exception, ):
        pass

    res['gkh_total_floors']= max_floor(res['gkh_total_floors'])

    if not pd.isna(res['building_page_url']):
        update_gkh_base(res['gkh_address'], res)
    else:
        print('Не получен ', res)
        logging.info(f'Объект {res} не получен')
    print(res)
    logging.info(f'Объект {res}')
    return res

# -
def renov_fill(ser):
    if str(ser[0]) != 'nan':
        return ser[0]
    elif ser[1] >= 2010:
        return 'новый дом'
    elif ser[1] < 2010:
        return 'нет в плане'
    else:
        return 'N/A'

# -
def max_floor(st):
    """ Функция очистки данных об этажности. Возвращает последнее число из строки и переводит в int"""
    try:
        return int(re.findall(r'\d+', st)[-1])
    except (Exception, ):
        return 0

# -
def qty_rooms_predict(ser):
    """ Функция предсказывает количествокомнат исходя из данных переданной серии pd.Series
        на обученном классификаторе дерева решений clf
        Возвращает серию"""
    global clf

    if pd.isna(ser.loc['qty_rooms']):
        try:
            fea_ = np.array(ser[['total_floors', 'construction_year', 'obj_square']]).reshape(1, -1)
            ser.loc['qty_rooms'] = clf.predict(fea_)[0]
        except Exception as e:
            ser.loc['qty_rooms'] = 0
    return ser

# -
def update_spreadsheet_hist(path: str, _df, startcol: int = 1, startrow: int = 1, sheet_name: str = "Прошедшие"):
    """Функция перезаписи данных в конкретный лист "Прошедшие".
    :param path: Путь до файла Excel
    :param _df: Датафрейм Pandas для записи
    :param startcol: Стартовая колонка в таблице листа Excel, куда буду писать данные
    :param startrow: Стартовая строка в таблице листа Excel, куда буду писать данные
    :param sheet_name: Имя листа в таблице Excel, куда буду писать данные
    :return: возвращает файл с датафреймом data_for_save (названия столбцов предустановлены),
            в ячейку А1 пишется дата создания файла"""
    global date_to
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    wb = ox.load_workbook(path)
    for ir in range(0, len(_df)):
        for ic in range(0, len(_df.iloc[ir])):
            wb[sheet_name].cell(startrow + ir, startcol + ic).value = _df.iloc[ir][ic]
            wb[sheet_name].cell(startrow + ir, startcol + ic).border = thin_border
            if ic in [5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 24]:
                wb[sheet_name].cell(startrow + ir, startcol + ic).alignment = Alignment(horizontal='center')
            elif ic in [15, 16, 17, 18]:
                wb[sheet_name].cell(startrow + ir, startcol + ic).alignment = Alignment(horizontal='center')
                wb[sheet_name].cell(startrow + ir, startcol + ic).number_format = r'# ##0'
            elif ic in [21]:
                wb[sheet_name].cell(startrow + ir, startcol + ic).alignment = Alignment(horizontal='center')
                wb[sheet_name].cell(startrow + ir, startcol + ic).number_format = 'DD.MM.YYYY'
            elif ic in [19]:
                wb[sheet_name].cell(startrow + ir, startcol + ic).alignment = Alignment(horizontal='center')
                wb[sheet_name].cell(startrow + ir, startcol + ic).number_format = '0.00%'
            elif ic in [25]:
                wb[sheet_name].cell(startrow + ir, startcol + ic).alignment = Alignment(horizontal='center')
                wb[sheet_name].cell(startrow + ir, startcol + ic).font = Font(color='808080')
                wb[sheet_name].cell(startrow + ir, startcol + ic).fill = PatternFill('solid', fgColor="D9D9D9")

    wb[sheet_name].cell(1, 1).value = pd.Timestamp('today').date() - pd.Timedelta(1, "d")
    wb[sheet_name].conditional_formatting = ConditionalFormattingList()
    cells_range = 'N2:N' + str(len(_df) + 1)
    pred_font = Font(color='0000FF')
    diff_style = DifferentialStyle(font=pred_font)
    rule = Rule(type="expression", dxf=diff_style)
    rule.formula = ["O2=1"]
    wb[sheet_name].conditional_formatting.add(cells_range, rule)
    wb.save(path)


def start_browser_for_parse(pict=False):
    """ Функция запуска Chrome-браузера"""

    chrome_options = Options()
    chrome_options.page_load_strategy = 'eager'

    #chrome_options = webdriver.ChromeOptions()
    # chrome_options.add_argument("--headless")
    # chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument('--start-maximized')
    chrome_options.add_argument('--log-level=3')
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    if not pict:
        prefs = {"profile.managed_default_content_settings.images": 2}
        chrome_options.add_experimental_option("prefs", prefs)
    path_to_wd = os.getcwd()
    # driver = webdriver.Chrome(executable_path=(path_to_wd + '\\chromedriver.exe'), options=chrome_options)
    service = webdriver.ChromeService(executable_path=(path_to_wd + '\\chromedriver.exe'))
    driver = webdriver.Chrome(service=service, options=chrome_options)
    return driver


def get_url_list_from_page(driver, cur_page_url, is_active):
    """ Функция получения списка URL всех лотов на странице cur_page_url
        При открытии страницы разворачивает все svg-объекты.
    :param driver: работающий в программе WebDriver
    :param cur_page_url: URL страницы investmoscow.ru для получения списка лотов
    :return: возвращает список URL
    """
    url_page_lst = []
    driver.get(cur_page_url)
    wait = WebDriverWait(driver, 20)
    a_class = (By.CLASS_NAME, "uid-mb-40")
    qty_a_class = len(driver.find_elements(By.CLASS_NAME, "uid-mb-40"))
    if qty_a_class > 0:
        wait.until(
            EC.presence_of_element_located(a_class))
    # ---> прокрутка до конца страницы
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    driver.execute_script("window.scrollTo(document.body.scrollHeight, 0);")
    # ---<
    time.sleep(2)
    arrows_down_qty = len(driver.find_elements(By.XPATH,
                                               '//div[@class="list"]/div/div[@class="uid-mb-40"]' +
                                               '/div[@class="uid-group-card__collapse"]/div[@class="uid-group-card-toggle"]/' +
                                               '/div/*[name()="svg"][@class="uid-arrow uid-arrow-down"]'))
    while arrows_down_qty > 0:
        #        print(arrows_down_qty)
        try:
            # wait.until(EC.presence_of_element_located((By.XPATH,
            driver.find_element(By.XPATH,
                                '//div[@class="list"]/div/div[@class="uid-mb-40"]' +
                                '/div[@class="uid-group-card__collapse"]/div[@class="uid-group-card-toggle"]/' +
                                '/div/*[name()="svg"][@class="uid-arrow uid-arrow-down"]').click()
            arrows_down_qty = len(driver.find_elements(By.XPATH,
                                                       '//div[@class="list"]/div/div[@class="uid-mb-40"]' +
                                                       '/div[@class="uid-group-card__collapse"]/div[@class="uid-group-card-toggle"]/' +
                                                       '/div/*[name()="svg"][@class="uid-arrow uid-arrow-down"]'))
        except (Exception,) as exc:
            pass

    # ------------------   конец - корректировка поиска стрелочек вниз
    wait.until(
        EC.presence_of_element_located((By.XPATH,
                                        '//div[@class="list"]')))
    body = driver.find_element(By.CSS_SELECTOR, 'body')

    # ----> медленная прокрутка окна
    driver.execute_script("window.scrollTo(document.body.scrollHeight, 0);")
    while True:
        # Scroll down to bottom
        body.send_keys(Keys.PAGE_DOWN)
        scroll_position = driver.execute_script("return window.scrollY;")
        max_position = driver.execute_script("return document.body.scrollHeight;")
        # Wait to load page
        time.sleep(SCROLL_PAUSE_TIME)
        # Calculate new scroll height and compare with last scroll height
        if scroll_position >= max_position - 979:
            print("break")
            break
    soup = BeautifulSoup(driver.find_element(By.XPATH, '//div[@class="list"]').get_attribute('innerHTML'),
                         "html.parser")
    nfo = soup.find_all('a', attrs={"class": "uid-tenders-card__main"})
    today = pd.to_datetime('today')
    for el in nfo:
        deadline = el.find('div', attrs={"class": "uid-text-small uid-text-gray"})\
                        .span.text.split(',')[0]
        deadline_date = datetime.datetime.strptime(deadline, "%d.%m.%Y").date()
        if (is_active and deadline_date>=today) or (not is_active and deadline_date<=today):
            url_page_lst.append('https://investmoscow.ru' + el['href'])

    return driver, url_page_lst


def reparsing_active_unreadable_urls(driver, lot_df, qty_lots, unreadable_urls):
    """ Функция повторного парсинга непрочитанных ранее активных лотов.
        :param lot_df: датафрейм с уже обработанными лотами
        :param qty_lots: количество всех лотов
        :param unreadable_urls: список URL необработанных лотов
        :return: датафрейм с обработанными лотами. Если в итоге не получается скачать некоторые лоты -
        выводит данные об этом.
        """
    for j in range(10):
        oper_lst = unreadable_urls.copy()
        if len(oper_lst) > 0:
            print(f'Попытка чтения списка номер {j + 2}')
            print(f'Нераспознанных ссылок {len(unreadable_urls)}')
            logging.info(f'Попытка чтения списка номер {j + 2}')
            logging.info(f'Нераспознанных ссылок {len(unreadable_urls)}')
            for url in tqdm(oper_lst):
                # считываем страницу с заданным URL
                try:
                    driver, lot_single = parse_lot(driver, url)
                    if not (pd.isna(lot_single['start_price'])):
                        lot_df = pd.concat([lot_df, lot_single.to_frame().T], ignore_index=True)
                    else:
                        print(f'Нечитаемая ссылка {url}')
                        print('lot_tag ' + lot_single['lot_tag'])
                        logging.info(f'Нечитаемая ссылка {url}')
                        logging.info(f"lot_tag  {lot_single['lot_tag']}")
                        continue
                    unreadable_urls.remove(url)
                except (Exception,):
                    print(f'Нечитаемая ссылка {url}')
                    logging.info(f'Нечитаемая ссылка {url}')
                    continue
        else:
            break
    if len(lot_df) == qty_lots:
        print('Все объекты скачаны')
        logging.info('Все объекты скачаны')
    else:
        print('Не удалось скачать объектов: ', len(unreadable_urls))
        logging.info(f'Не удалось скачать объектов: {len(unreadable_urls)}')
        print(unreadable_urls)
        logging.info(f'Объекты {unreadable_urls}')
    return lot_df


def reparsing_historical_unreadable_urls(driver, hist_df, bidding_df,
                                         final_df, unreadable_urls, today):
    """ Функция повторного парсинга непрочитанных ранее прошедших лотов.
        :param lot_df: датафрейм с уже обработанными лотами
        :param qty_lots: количество всех лотов
        :param unreadable_urls: список URL необработанных лотов
        :return: датафрейм с обработанными лотами. Если в итоге не получается скачать некоторые лоты -
        выводит данные об этом.
        """

    for j in range(10):
        oper_lst = unreadable_urls.copy()
        if len(oper_lst) > 0:
            print(f'Попытка чтения списка номер {j + 2}')
            print(f'Нераспознанных ссылок {len(unreadable_urls)}')
            logging.info(f'Попытка чтения списка номер {j + 2}')
            logging.info(f'Нераспознанных ссылок {len(unreadable_urls)}')
            for url in tqdm(oper_lst):
                # считываем страницу с заданным URL
                try:
                    lot_status_torgi = 0
                    if url not in final_df['investmoscow_url']:
                        driver, lot_single = parse_lot(driver, url)

                    if pd.to_datetime(lot_single['results_date']) > today - datetime.timedelta(days=7):
                        continue
                    driver, lot_status_data = control_lot(driver, url)

                    status_lst = ['ПРИЗНАНЫ НЕСОСТОЯВШИМИСЯ', 'ПРИЗНАНЫ СОСТОЯВШИМИСЯ', 'ОТМЕНЕНЫ']
                    try:
                        lot_reestr_num = lot_single['roseltorg_url'].split('/')[-1]
                    except:
                        print(url, lot_single['roseltorg_url'])
                        lot_reestr_num = str(lot_single['roseltorg_url']).split('/')[-1]
                    if lot_reestr_num in bidding_df.reestr_num.values:
                        #           if bidding_df[bidding_df.reestr_num==lot_reestr_num].loc['reestr_num']:
                        lot_bidding_data = bidding_df.loc[bidding_df.reestr_num == lot_reestr_num]
                        if lot_status_data['status'] not in status_lst:
                            try:
                                if (lot_bidding_data.iloc[0, 3] > 0) or (lot_bidding_data.iloc[0, 2] == 1):
                                    lot_status_data['status'] = 'ПРИЗНАНЫ СОСТОЯВШИМИСЯ'
                                    if lot_bidding_data.iloc[0, 3] > 0:
                                        lot_status_data['final_price'] = lot_bidding_data.iloc[0, 3]
                                    else:
                                        lot_status_data['final_price'] = lot_single['start_price']
                            except:
                                continue
                        row_to_add = pd.concat([lot_single, pd.Series(lot_status_data)])
                        try:
                            row_to_add['final_price_m2'] = round(row_to_add['final_price'] / row_to_add['obj_square'],
                                                                 2)
                            row_to_add['delta_price'] = round(
                                (row_to_add['final_price'] - row_to_add['start_price']) / row_to_add['start_price'], 4)
                        except (Exception,):
                            pass
                        hist_df = pd.concat([hist_df, row_to_add.to_frame().T], ignore_index=True)
                    else:
                        print(lot_reestr_num, bidding_df[bidding_df.reestr_num == lot_reestr_num])
                        continue
                    unreadable_urls.remove(url)
                except (Exception,):
                    print(f'Нечитаемая ссылка {url}')
                    logging.info(f'Нечитаемая ссылка {url}')
                    continue
        else:
            break
    if len(unreadable_urls) == 0:
        print('Все объекты скачаны')
        logging.info('Все объекты скачаны')
    else:
        print('Не удалось скачать объектов: ', len(unreadable_urls))
        logging.info(f'Не удалось скачать объектов: {len(unreadable_urls)}')
        print(unreadable_urls)
        logging.info(f'Объекты {unreadable_urls}')
    return hist_df


def drop_duplicated_lots(lot_df):
    """ Функция чистки датафрейма от дублирующих лотов по lot_tag. (Остаются последние по дате аукциона)
        :param lot_df: полный датафрейм после парсинга лотов
        :return: датафрейм с уникальными лотами. Выводит данные о количестве уникальных лотов.
    """
    # поиск дублирующихся лотов

    last_dup = lot_df[lot_df.duplicated(['lot_tag'], keep=False)]
#    last_dup.bidding_date.apply(lambda x: print(x, isinstance(x, pd.Timestamp)))
    last_dup.bidding_date = last_dup.apply(lambda x: x.bidding_date
                                                        if isinstance(x.bidding_date, pd.Timestamp)
                                                        else pd.Timestamp(x.bidding_date), axis=1)
    last_dup.bidding_date = last_dup.apply(lambda x: x.bidding_date
                                                                    if x.bidding_date != 'Нет данных'
                                                                    else pd.Timestamp('2000-01-01'),
                                           axis=1)
    last_dup = last_dup.sort_values('bidding_date')
    last_dup = last_dup.drop_duplicates(subset=['lot_tag'], keep='last')
    # удаляем дублирующиеся лоты по тагу
    lot_df = lot_df.drop_duplicates(subset=['lot_tag'], keep=False, ignore_index=True)
    # и дописываем оптимальные из дубликатов к датафрейму
    lot_df = lot_df.append(last_dup).reset_index(drop=True)
    print('Количество уникальных объектов: ', len(lot_df))
    logging.info(f'Количество уникальных объектов: {len(lot_df)}')
    return lot_df

# -
def compare_with_final_df(lot_df, final_df):
    lot_df = lot_df.merge(final_df[['lot_tag', 'addr_norm']].rename(columns={'addr_norm':'is_present'}),
                          how='left', on='lot_tag')
    lot_df = lot_df[lot_df.is_present.isna()]
    lot_df = lot_df.drop(columns=['is_present'])
    return lot_df

# -
def recognize_and_normalize_addresses(norm_df):
    """ Функция нормализации адресов датафрейма (приведения к единому стандарту).
        Сравниваются адреса, полученные с помощью NER-модуля Natasha из объявлений, с адресами в базе ЖКХ.
        При отсутствии совпадений - провдоится повтоная нормализация с использованием словаря преобразований
        conv_dictionary.xlsx. Если совпадений нет и адрес распознан нормально (определяется путем сверки
        строки адреса с преобразованным адресом визуально) - можно добавить данные вручную в базу ЖКХ.
        :param norm_df: полный датафрейм после парсинга лотов
        :return: norm_df: датафрейм с базой ЖКХ и нормализованными адресами.
    """
    global gkh_df
    # считываем базу адресов с указанием года постройки и ссылки на дом (далее - база ЖКХ)
    print('Считывается база ЖКХ')
    logging.info('Считывается база ЖКХ')
    gkh_df = pd.read_csv(GKH_BASE_FILENAME)
    gkh_df['gkh_address'] = gkh_df.gkh_address.apply(lambda x: str(x).lower())

    # проводим первичную нормализацию адресов датафрейма
    print('Первичная нормализация адресов датафрейма')
    logging.info('Первичная нормализация адресов датафрейма')
    norm_df.loc[:, 'addr_street'], norm_df.loc[:, 'addr_build_num'], \
        norm_df.loc[:, 'addr_floor_from_title'], norm_df.loc[:, 'addr_apart_num_from_title'] = zip(
            *norm_df.loc[:, 'addr_string'].progress_apply(primary_addr_normalize))

    # добавляем колонки с нормализованным адресом, и приведенным к нижнему регистру
    # по приведенному к нижнему регистру адресу пытаемся сджойнить с базой ЖКХ.
    norm_df['addr_norm'] = norm_df['addr_street'] + norm_df['addr_build_num']
    norm_df['addr_norm_lower'] = norm_df['addr_norm'].apply(lambda x: str(x).lower())
    norm_df = norm_df.merge(gkh_df, how='left', left_on='addr_norm_lower', right_on='gkh_address')

    # выделяем из базы строки, для которых не нашлось соответствия в базе ЖКХ
    quest_df = norm_df[norm_df.gkh_address.isna()]
    norm_df = norm_df.drop(norm_df[norm_df.gkh_address.isna()].index)

    # создаем цикл обработки необработанных адресов
    if len(quest_df) > 0:
        loop_flag = True
    else:
        loop_flag = False
    while loop_flag:
        # если обновлялась база ЖКХ
        # считываем базу адресов с указанием года постройки и ссылки на дом (далее - база ЖКХ)
        print('Повторная нормализация необработанных ранее адресов датафрейма')
        logging.info('Повторная нормализация необработанных ранее адресов датафрейма')
        print('Считывается база ЖКХ')
        gkh_df = pd.read_csv(GKH_BASE_FILENAME)
        gkh_df['gkh_address'] = gkh_df.gkh_address.apply(lambda x: str(x).lower())
        # убираем из нее ранее добавленные данные из ЖКХ (чтобы потом повторно попробовать сджойнить)
        quest_df = quest_df.drop(columns=GKH_FIELDS)
        # проводим повторную нормализацию адресов датафрейма
        quest_df.loc[:, 'addr_street'], quest_df.loc[:, 'addr_build_num'], \
            quest_df.loc[:, 'addr_floor_from_title'], quest_df.loc[:, 'addr_apart_num_from_title'] \
            = zip(*quest_df.loc[:, 'addr_string'].progress_apply(secondary_addr_normalize))
        # добавляем колонки с нормализованным адресом, и приведенным к нижнему регистру
        # по приведенному к нижнему регистру адресу пытаемся сджойнить с базой ЖКХ.
        quest_df['addr_norm'] = quest_df['addr_street'] + quest_df['addr_build_num']
        quest_df['addr_norm_lower'] = quest_df['addr_norm'].apply(lambda x: str(x).lower())
        quest_df = quest_df.merge(gkh_df, how='left', left_on='addr_norm_lower', right_on='gkh_address')
        # те строки, которые получилось сджойнить, добавляем в norm_df
        norm_df = norm_df.append(quest_df.query('addr_norm_lower == gkh_address'))
        quest_df = quest_df.query('addr_norm_lower != gkh_address')
        if len(quest_df) == 0:
            break
        else:
            print('Нераспознанных адресов: ', len(quest_df))
            logging.info(f'Нераспознанных адресов: {len(quest_df)}')
            # новая версия - выдает на печать только по одной паре реальный адрес / распознанный
            quest_df = quest_df.reset_index(drop=True)
            quest_df_sample = quest_df.drop_duplicates(subset=['addr_norm_lower']).copy()
            for i, row in quest_df_sample.iterrows():
                try:
                    print(row['addr_string'])
                    print(row['addr_norm_lower'])
                    print('Если нормализованный адрес совпадает с адресом в строке объявления - нажмите Y')
                    is_cont = input()
                    while is_cont not in ['Y', 'y', 'N', 'n']:
                        print('Некорректный ввод. Повторите выбор (Y/N)')
                        is_cont = input()
                    if is_cont in ['Y', 'y']:
                        new_addr_dict = {'gkh_address': row['addr_norm_lower']}
                        gkh_df = pd.concat([gkh_df, pd.DataFrame.from_records([new_addr_dict])] ,ignore_index=True)
                except Exception as e:
                    logging.info(f'Ошибка: {traceback.format_exc().split("Stacktrace:")[0]}')
            gkh_df.to_csv(GKH_BASE_FILENAME, index=False)


            # TODO: откорректировать поиск нераспознанных
#                add_unrecognized()
            quest_df = quest_df.drop(columns=GKH_FIELDS)
            quest_df['addr_norm_lower'] = quest_df['addr_norm'].apply(lambda x: str(x).lower())
            quest_df = quest_df.merge(gkh_df, how='left', left_on='addr_norm_lower', right_on='gkh_address')
            norm_df = norm_df.append(quest_df.query('addr_norm_lower == gkh_address'))
            quest_df = quest_df.query('addr_norm_lower != gkh_address')
            if len(quest_df) == 0:
                print('Все адреса распознаны')
                logging.info('Все адреса распознаны')
                loop_flag = False
            else:
                print('Продолжить попытки распознавания адресов? (Y/N')
                is_cont = input()
                while is_cont not in ['Y', 'y', 'N', 'n']:
                    print('Некорректный ввод. Повторите выбор (Y/N)')
                    is_cont = input()
                if is_cont in ['N', 'n']:
                    loop_flag = False

    return norm_df

def control_new_gkh_df_2gis():
    """Функция для сравнения данных в new_gkh_df с данными на 2gis. Сравнивает дату постройки дома
    и этажность для каждого адреса. При совпадении - фиксирует данные о доме в общей базе ЖКХ.
    Несовпадающие оставляет в new_gkh_df"""
    global gkh_df
    global new_gkh_df

    driver_2gis = start_browser_for_parse(pict=True)
    new_gkh_df = new_gkh_df.reset_index(drop=True)
    for i, row in tqdm(new_gkh_df.iterrows(), position=1):
        try:
            cont_year_2gis = 0
            tot_floor_2gis = 0
            is_controlled = False
            driver_2gis.get('https://2gis.ru/moscow')
            element = driver_2gis.find_element(By.XPATH, "//input[@class='_1gvu1zk']")
            adr1 = row['gkh_address']
            if adr1.startswith('г. москва, г.'):
                re.sub("г. москва, г. ", "г. ", adr1)
            element.send_keys(adr1)
            element.send_keys(Keys.ENTER)
            time.sleep(2)
            driver_2gis.find_element(By.XPATH, "//div[@class='_zjunba']").click()
            time.sleep(2)
            fl_el = driver_2gis.find_element(By.XPATH, "//div[@class='_49kxlr']/span/span[2]")
            try:
                tot_floor_2gis = int(fl_el.text.split(' ')[0])
                if tot_floor_2gis != row['gkh_total_floors']:
                    # print('2 gis этажность - ', tot_floor_2gis)
                    # print('flatinfo этажность - ', row['gkh_total_floors'])
                    is_controlled = True
            except Exception as exc:
                # print(' 2 gis этажность неопределена' )
                is_controlled = True
            table_elements = driver_2gis.find_elements(By.XPATH, "//li[@class='_4rm1c']")
            for el in table_elements:
                if el.text.split('\n')[0] == 'Год постройки':
                    try:
                        cont_year_2gis = int(el.text.split('\n')[1])
                        if abs(cont_year_2gis - row['construction_year']) > 1:
                            is_controlled = True
                            # print('2 gis год постройки - ', cont_year_2gis)
                            # print('flatinfo год постройки - ', row['construction_year'])
                    except:
                        print(' 2 gis год постройки не определен')
                        is_controlled = True
                #print(el.find_element(By.XPATH, "//span[@class='_18zamfw']").text)

        except Exception as exc:
            print(exc)
            is_controlled = True

        if not is_controlled:
            gkh_df = pd.concat([gkh_df, row.to_frame().T])
            new_gkh_df = new_gkh_df.drop(i)
        if is_controlled:
            print(row['gkh_address'])
            print('2 gis этажность - ', tot_floor_2gis)
            print('flatinfo этажность - ', row['gkh_total_floors'])
            print('2 gis год постройки - ', cont_year_2gis)
            print('flatinfo год постройки - ', row['construction_year'])
            print(is_controlled)
    driver_2gis.quit()
    if len(new_gkh_df) > 0:
        update_new_buildings(r'..\realty_model_files\new_buildings.xlsx', new_gkh_df,
                             startcol=1, startrow=2)
        print('Не все данные распознались успешно')
        logging.info('Не все данные распознались успешно')
        print('Внесите данные о метро и этажности вручную в файл new_buildings.xlsx.')
        print('Затем сохраните и закройте файл. Нажмите Enter.')
        input()
        add_unrecognized()
    else:
        gkh_df = gkh_df.drop_duplicates(subset=['gkh_address'], keep='last')
        gkh_df.to_csv(GKH_BASE_FILENAME, index=False)

def check_out_of_data_metro_floor(norm_df):
    """ Функция проверки полноты данных о привязке к метро и этажности зданий, году постройки. При отсутствии данных в базе ЖКХ -
            запускается попытка скачать с сайта FlatInfo.ru, при невозможности - дает возможность доплнения данных
            базы ЖКХ вручную (через new_buildings.xlsx)
            :param norm_df: полный датафрейм после парсинга лотов
            :return: norm_df: датафрейм с базой ЖКХ и нормализованными адресами.
        """
    global driver
    global new_gkh_df
    global gkh_df

    out_of_data = norm_df[
        norm_df.gkh_metro_station.isna() | norm_df.gkh_total_floors.isna() | norm_df.construction_year.isna()] \
        .drop_duplicates(subset=['addr_norm'], keep='last')
    if out_of_data.shape[0] != 0:
        #driver = start_browser_for_parse()
        print('Отсутвуют данные по ', len(out_of_data), ' объектам')
        logging.info(f'Отсутвуют данные по {len(out_of_data)} объектам')
        print('Дождитесь окончания сбора и сохранения данных')
        out_of_data.progress_apply(lambda x: metro_and_floor_data(x.gkh_address, x.building_page_url), axis=1)
        #driver.quit()
        # проверка new_gkh_df через 2gis
        control_new_gkh_df_2gis()
        gkh_df = pd.read_csv(GKH_BASE_FILENAME)
        gkh_df['gkh_address'] = gkh_df.gkh_address.apply(lambda x: str(x).lower())
        print('Данные сохранены')
        logging.info('Данные сохранены')
        norm_df = norm_df.drop(columns=GKH_FIELDS)
        norm_df = norm_df.merge(gkh_df, how='left', left_on='addr_norm_lower', right_on='gkh_address')
    return norm_df

def is_metro_and_floor_data_complete(norm_df):
    """ Функция, запускающая проверки полноты данных о привязке к метро и этажности зданий,
        году постройки. При отсутствии данных в базе ЖКХ -
        запускается попытка скачать с сайта FlatInfo.ru, при невозможности - дает возможность доплнения данных
        базы ЖКХ вручную (через new_buildings.xlsx)
        :param norm_df: полный датафрейм после парсинга лотов
        :return: norm_df: датафрейм с базой ЖКХ и нормализованными адресами.
    """
    print('Проверка наличия данных о метро и этажности зданий')
    logging.info('Проверка наличия данных о метро и этажности зданий')
    norm_df = check_out_of_data_metro_floor(norm_df)
    norm_df = check_out_of_data_metro_floor(norm_df)

    return norm_df

# -
def fill_spaces_in_data(norm_df, is_for_winner = False):
    """ Функция заполняет пропуски в данных, добавляет информацией о реновации,
        добавляет timestamp для сортировки.
        :param norm_df: полный датафрейм после парсинга лотов
        :return: norm_df: датафрейм с базой ЖКХ и нормализованными адресами.
    """
    # заполняем пропуски данными из подтянутых баз
    norm_df = norm_df.reset_index(drop=True)
    norm_df['total_floors'] = norm_df[['total_floors', 'gkh_total_floors']].max(axis=1)
    def fill_cadastr_num(x):
        if x.cadastr_num == "Не указано" or pd.isna(x.cadastr_num):
            if not pd.isna(x.gkh_cadastr_num):
                return x.gkh_cadastr_num
            else:
                return "Нет данных"
        else:
            return x.cadastr_num
    if not is_for_winner:
        norm_df['cadastr_num'] = norm_df.apply(lambda x: fill_cadastr_num(x), axis=1)
        norm_df['renov_period'] = norm_df['renov_period'].fillna("Не включен в программу")
        norm_df['addr_floor'] = norm_df['addr_floor'].fillna(norm_df['addr_floor_from_title']).fillna("Нет данных")
        norm_df['flat_num'] = norm_df['flat_num'].fillna(norm_df['addr_apart_num_from_title']).fillna("Нет данных")
        if norm_df['qty_rooms'].isna().sum() > 0:
            norm_df = fill_qty_rooms_with_predictions(norm_df)
        NUMERIC_COLUMNS = ['inf_sales', 'inf_food_service', 'inf_education',
                           'inf_cult_and_sport', 'inf_consumer_services', 'inf_health_care',
                           'residents_qty', 'ceiling_height', 'passenger_elevators_qty']
    else:
        NUMERIC_COLUMNS = ['residents_qty', 'ceiling_height', 'passenger_elevators_qty']

    norm_df['is_total_floors_variable'] = norm_df['is_total_floors_variable'].fillna(0)
    norm_df[NUMERIC_COLUMNS] = norm_df[NUMERIC_COLUMNS].fillna(-1)
    norm_df = norm_df.fillna("Нет данных")

    return norm_df

# -
def fill_qty_rooms_with_predictions(norm_df):
    """ Функция заполняет пропуски в данных о количестве комнат с ипользованием ML-модели (DecisionTree),
        обучаемой на датасете прошедших торгов с известными параметрами
        (кол-во комнат <--- ujl постройки дома, этажность, площадь объекта).
        :param norm_df: полный датафрейм после парсинга лотов
        :return: norm_df: датафрейм с базой ЖКХ и нормализованными адресами.
    """
    global clf
    # скачиваем исторические данные
    df1 = pd.read_csv('historical_data.csv')
    df1 = df1[~df1.qty_rooms.isna()]
    df1 = df1[df1.is_qty_rooms_predicted == 0]
    df1 = df1[['total_floors', 'year_exp_w', 'qty_rooms', 'obj_square']]
    df1.total_floors = df1.total_floors.apply(lambda x: max_floor(str(x)))
    # и обучаем на них дерево решений
    clf = tree.DecisionTreeClassifier(criterion='entropy', max_depth=10)
    X = df1[['total_floors', 'year_exp_w', 'obj_square']]
    y = df1[['qty_rooms']]
    clf.fit(X, y)
    # чистим данные об этажности
    # и пронозируем количество комнат
    norm_df = norm_df.apply(lambda x: qty_rooms_predict(x), axis=1)
    return norm_df

# -
def compare_with_existing(ser):
    """ Функция сравнения Series, переданной из нового датафрейма, со строками в скачанном файле
        MosInvestData.xlsx.
        Пытается найти номер переданного лота в существующем датафрейме.
        При отстутсвии возвращает -1 в поле flag.
        При наличии в датафрейме - сравнивает данные, отличающиеся данные заменяются на новые,
        в поле "Примечания" вписывается дата внесения изменений, flag = -2
        Если данные корректны - возвращает номер строки датафрейма
    """
    global MID_df
    today = dt.datetime.today().date()
    try:
        row = MID_df.loc[MID_df.lot_tag == ser.lot_tag].index[0]
    except (Exception, ):
        return -1
    if str(MID_df.loc[row, 'deadline']) != str(ser.loc['deadline']):
        MID_df.loc[row, 'deadline'] = ser.loc['deadline']
        MID_df.loc[row, 'Примечания'] = today
        return -2
    if str(MID_df.loc[row, 'auct_date']) != str(ser.loc['auct_date']):
        MID_df.loc[row, 'auct_date'] = ser.loc['auct_date']
        MID_df.loc[row, 'Примечания'] = today
        return -2
    return row

# -
def metro_prep(name):
    """ Функция перевода имени метро в нижний регистр и уточнения отдельных станций"""
    try:
        if name.lower() == 'покровская':
            name = 'покровское'
        elif name.lower() == 'кубанская (люблино)':
            name = 'люблино'
        elif name.lower() == 'новые черёмушки':
            name = 'новые черемушки'
        name = name.lower()
    except (Exception,):
        pass
    return name

# -
def copy_cell(src_sheet, src_row, src_col,
              tgt_sheet, tgt_row, tgt_col,
              copy_style=True):
    """Функция копирования свойств ячейки с одного листа на другой вместе с содержимым """
    cell = src_sheet.cell(src_row, src_col)
    new_cell = tgt_sheet.cell(tgt_row, tgt_col, cell.value)
    if cell.has_style and copy_style:
        new_cell._style = copy(cell._style)

# -
def update_spreadsheet_w_drop(path: str, _df, startcol: int = 1, startrow: int = 1, sheet_name: str = "Sheet1",
                              add_rows_qty: int = 0):
    """
    Функция перезаписи данных в конкретный лист с формартированием - применимо для шаблонов листов "Активные"
    и "NEW!!!". Копирует в лист "избранное" строки, помеченные заливкой на листе "Активные", подлежащие удалению
    в связи с истечением срока (дата аукциона+2дня -> flag = -2). Устанавливает правила условного форматирования
    данных. Если в строке flag=False (поле 30) --> строка удаляется
    :param path: Путь до файла Excel
    :param _df: Датафрейм Pandas для записи
    :param startcol: Стартовая колонка в таблице листа Excel, куда буду писать данные
    :param startrow: Стартовая строка в таблице листа Excel, куда буду писать данные
    :param sheet_name: Имя листа в таблице Excel, куда буду писать данные
    :param add_rows_qty: Количество строк для добавления выше последней
    :return: возвращает файл с датафреймом _df (названия столбцов предустановлены),
            в ячейку А1 пишется дата создания файла

    """
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    today = pd.Timestamp('today').date()
    file_name = path
    wb = ox.load_workbook(path)
    new_row_fav = wb['Избранное'].max_row + 1
    cnt_fav = 0
    if add_rows_qty > 0:
        wb[sheet_name].insert_rows(len(_df) + startrow - add_rows_qty, add_rows_qty)
    for r_ in range(add_rows_qty):
        for c_ in range(25):
            wb[sheet_name].cell(row=len(_df) + startrow - r_ - 1, column=c_ + 1).border = thin_border
    # перебираем скачанный датафрейм с добавленными строками с конца
    # (последняя строка остается - символ конца файла)
    for ir in range(len(_df) - 1, -1, -1):
        # проверка состояния флага (True - записать, False - удалить)
        #        print(ir, _df.iloc[ir][26])
        if _df.iloc[ir][26]:

            for ic in range(0, len(_df.iloc[ir]) - 2):
                wb[sheet_name].cell(startrow + ir, startcol + ic).value = _df.iloc[ir][ic]
                if ic in [4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 22]:
                    wb[sheet_name].cell(startrow + ir, startcol + ic).alignment = Alignment(horizontal='center')
                elif ic in [15, 16]:
                    wb[sheet_name].cell(startrow + ir, startcol + ic).alignment = Alignment(horizontal='center')
                    wb[sheet_name].cell(startrow + ir, startcol + ic).number_format = "# ##0"
                elif ic in [18, 19]:
                    wb[sheet_name].cell(startrow + ir, startcol + ic).alignment = Alignment(horizontal='center')
                    wb[sheet_name].cell(startrow + ir, startcol + ic).number_format = 'DD.MM.YYYY'
                elif ic in [23]:
                    wb[sheet_name].cell(startrow + ir, startcol + ic).alignment = Alignment(horizontal='center')
                    wb[sheet_name].cell(startrow + ir, startcol + ic).font = Font(color='808080')
                    wb[sheet_name].cell(startrow + ir, startcol + ic).fill = PatternFill('solid', fgColor="D9D9D9")
        else:
            if wb[sheet_name].cell(startrow + ir, 1).fill.start_color.index != '00000000':
                for i, row in enumerate(wb[sheet_name].iter_rows(min_row=startrow + ir, max_row=startrow + ir), 1):
                    for cell in row:
                        copy_cell(wb[sheet_name], cell.row, cell.column,
                                  wb['Избранное'], new_row_fav + cnt_fav, cell.column)
                cnt_fav += 1

            wb[sheet_name].delete_rows(startrow + ir, 1)
    if cnt_fav > 0:
        wb['Избранное'].conditional_formatting = ConditionalFormattingList()
    wb[sheet_name].cell(1, 1).value = today
    # форматирование цвета цены
    color_scale_rule = ColorScaleRule(start_type="min",
                                      start_color="57FFA3",  # зеленый
                                      end_type="max",
                                      end_color="FFEF9C")  # желтый

    # добавим этот градиент снова к столбцу `Q`.
    if len(_df) > 0:
        wb[sheet_name].conditional_formatting = ConditionalFormattingList()
        cells_range = 'R2:R' + str(_df.flag.sum() + 1)
        if (_df.flag.sum() + 1) >2:
            wb[sheet_name].conditional_formatting.add(cells_range, color_scale_rule)
        # форматирование цвета дат
            date_cells_range = 'T2:U' + str(_df.flag.sum() + 1)
        # redFill = PatternFill(fill_type='solid', start_color='FECECE', end_color='FECECE')
        # greenFill = PatternFill(fill_type='solid', start_color='D8E4BC', end_color='D8E4BC')
        # blueFill = PatternFill(fill_type='solid', start_color='B7DEE8', end_color='B7DEE8')
            redFill = PatternFill(fill_type='solid', bgColor='FECECE')
            greenFill = PatternFill(fill_type='solid', bgColor='D8E4BC')
            blueFill = PatternFill(fill_type='solid', bgColor='B7DEE8')
            diff_style_blue = DifferentialStyle(fill=blueFill)
            rule_blue = Rule(type="expression", dxf=diff_style_blue)
            rule_blue.formula = ["(T2-$A$1)<=0"]
            diff_style_green = DifferentialStyle(fill=greenFill)
            rule_green = Rule(type="expression", dxf=diff_style_green)
            rule_green.formula = ["(T2-$A$1)>7"]
            diff_style_red = DifferentialStyle(fill=redFill)
            rule_red = Rule(type="expression", dxf=diff_style_red)
            # rule_red.formula = ["=AND((T2-$A$1)<=7,(T2-$A$1)>0)"]
            rule_red.formula = ["(T2-$A$1)<=7"]

            wb[sheet_name].conditional_formatting.add(date_cells_range, rule_blue)
            wb[sheet_name].conditional_formatting.add(date_cells_range, rule_green)
            wb[sheet_name].conditional_formatting.add(date_cells_range, rule_red)

            cells_range = 'N2:N' + str(_df.flag.sum() + 1)
            pred_font = Font(color='0000FF')
            diff_style = DifferentialStyle(font=pred_font)
            rule = Rule(type="expression", dxf=diff_style)
            rule.formula = ["O2=1"]
            wb[sheet_name].conditional_formatting.add(cells_range, rule)
    wb.save(file_name)

def add_bidding_data(norm_df, bidding_df):
    # bidding_df = pd.read_csv(BIDDING_FILENAME)
    # bidding_df.id = bidding_df.id.apply(str)
    # bidding_df.reestr_num = bidding_df.apply(lambda x: x.reestr_num.strip("'").strip()
    #                                                     if not x.reestr_num.startswith("'178")
    #                                                     else str(x.id), axis=1)
    norm_df = norm_df.merge(bidding_df, how='left', left_on='reestr_num_lot', right_on='reestr_num')
    print('Roseltorg: Отсутствуют данные по ', norm_df.id.isna().sum(), 'лотам')
    return norm_df


def historical_processing():
    global driver
    global gkh_df
    global current_date_string

    # считываем данные о прошедших торгах из файла
    print('Номер страницы')
    num = input()
    final_df = pd.read_csv(LOT_FILENAME_HISTORICAL)
 #   driver = start_browser_for_parse()
    driver.implicitly_wait(10)
    # url до 26.04.2023 'https://investmoscow.ru/tenders?pageNumber=1&pageSize=100&orderBy=CreateDate&orderAsc=false&objectTypes=7&tenderTypes=13&tenderStatuses=1&tradeForms=45001'
    main_page_url = 'https://investmoscow.ru/tenders?pageNumber=1&pageSize=10&orderBy=RequestEndDate&orderAsc=false&objectTypes=nsi:41:30011568&objectKinds=nsi:tender_type_portal:13&tenderStatus=nsi:tender_status_tender_filter:2&timeToPublicTransportStop.noMatter=true'
    driver.get(main_page_url)
    wait = WebDriverWait(driver, 10)
    if re.search('-',num):
        start_page = int(num.split('-')[0])
        end_page = int(num.split('-')[1])
    else:
        start_page = int(num)
        end_page = start_page
    url_lst = []
    for page_num in trange(start_page, end_page+1):
        cur_page_url = 'https://investmoscow.ru/tenders?pageNumber=' + str(page_num) + '&pageSize=100&' + \
                   'orderBy=RequestEndDate&orderAsc=false&objectTypes=nsi:41:30011568&' + \
                   'objectKinds=nsi:tender_type_portal:13&tenderStatus=nsi:tender_status_tender_filter:2' + \
                   '&timeToPublicTransportStop.noMatter=true'

        driver, url_page_lst = get_url_list_from_page(driver, cur_page_url, is_active=False)
        url_lst = url_lst + url_page_lst
    hist_df = pd.DataFrame()
    unreadable_urls = []
    print('Обработка страницы ', num, 'прошедших торгов')
    logging.info(f'Обработка страницы {num} прошедших торгов ')
    today = pd.to_datetime('today')

    bidding_df = pd.read_csv(BIDDING_FILENAME)
    bidding_df.id = bidding_df.id.apply(str)
    bidding_df.reestr_num = bidding_df.apply(lambda x: x.reestr_num.strip("'").strip()
                                                        if not x.reestr_num.startswith("'178")
                                                        else str(x.id), axis=1)

    for url in tqdm(url_lst, position=0):
        # считываем страницу с заданным URL, при нескачивании в течении 10 секунд добавляем ссылку в список нескачанных
        try:

            lot_status_torgi = 0
            if url not in final_df['investmoscow_url']:
                driver, lot_single = parse_lot(driver, url)
# если дата подведения результатов не наступила - переходим к след. url
            if pd.to_datetime(lot_single['results_date']) >= today: #- datetime.timedelta(days=7):
                continue
            driver, lot_status_data = control_lot(driver, url)

            status_lst = ['ПРИЗНАНЫ НЕСОСТОЯВШИМИСЯ', 'ПРИЗНАНЫ СОСТОЯВШИМИСЯ', 'ОТМЕНЕНЫ']
            try:
                lot_reestr_num = lot_single['roseltorg_url'].split('/')[-1]
            except:
                print(url, lot_single['roseltorg_url'])
                lot_reestr_num = str(lot_single['roseltorg_url']).split('/')[-1]
            if lot_reestr_num in bidding_df.reestr_num.values:
                lot_bidding_data = bidding_df.loc[bidding_df.reestr_num == lot_reestr_num]
                if lot_status_data['status'] not in status_lst:
                    try:
                        if (lot_bidding_data.iloc[0,3] > 0) or (lot_bidding_data.iloc[0,2] == 1):
                            lot_status_data['status'] = 'ПРИЗНАНЫ СОСТОЯВШИМИСЯ'
                            if lot_bidding_data.iloc[0,3] > 0:
                                lot_status_data['final_price'] = lot_bidding_data.iloc[0,3]
                            else:
                                lot_status_data['final_price'] = lot_single['start_price']
                    except:
                        continue
                row_to_add = pd.concat([lot_single, pd.Series(lot_status_data)])
                try:
                    row_to_add['final_price_m2'] = round(row_to_add['final_price'] / row_to_add['obj_square'], 2)
                    row_to_add['delta_price'] = round(
                        (row_to_add['final_price'] - row_to_add['start_price']) / row_to_add['start_price'], 4)
                except (Exception,):
                    pass
                hist_df = pd.concat([hist_df, row_to_add.to_frame().T], ignore_index=True)
            else:
                print(lot_reestr_num, bidding_df[bidding_df.reestr_num == lot_reestr_num])
                continue

        except Exception as e:
            print(f'Нечитаемая ссылка {url}')
            logging.info(f'Нечитаемая ссылка {url}')
            unreadable_urls.append(url)
            print('Ошибка:\n', traceback.format_exc().split('Stacktrace:')[0])
            logging.info(f'Ошибка:\n {traceback.format_exc().split("Stacktrace:")[0]}')
            continue

    print('Пропущенных объектов: ', len(unreadable_urls))
    logging.info(f'Пропущенных объектов: {len(unreadable_urls)}')
    if len(unreadable_urls) > 0:
        hist_df = reparsing_historical_unreadable_urls(driver, hist_df, bidding_df,
                                                       final_df, unreadable_urls, today)
    if len(hist_df) > 0:
        hist_df = drop_duplicated_lots(hist_df)
    else:
        print('Новые данные отсутвуют')
        logging.info('Новые данные отсутвуют')
    # нормируем и распознаем адреса (если есть новые)
    if len(hist_df) > 0:
        hist_df = recognize_and_normalize_addresses(hist_df)
        # после распознавания адресов - уточняем данные о метро и этажности
        hist_df = is_metro_and_floor_data_complete(hist_df)
        hist_df = fill_spaces_in_data(hist_df)
    else:
        hist_df = pd.DataFrame()
    hist_cols = OUTPUT_ACTIVE_COLS + HISTORICAL_LOT_FIELDS
    final_df = pd.concat([final_df, hist_df], ignore_index=True)
    final_df = drop_duplicated_lots(final_df)
    final_df[hist_cols].to_csv(LOT_FILENAME_HISTORICAL, index=False)
# ----> добавляем информацию о торгах с roseltorg.ru
    final_df['reestr_num_lot'] = final_df.roseltorg_url.apply(lambda x: x.split('/')[-1])
    final_df = add_bidding_data(final_df, bidding_df)
    final_df.drop(columns=['id', 'reestr_num'], inplace=True)
    output_columns = hist_cols + ['participant_qty', 'win_price', 'win_name']
    final_df[output_columns].to_csv("..\\realty_model_files\\output\\"
                                        + current_date_string[:8] + '_'
                                        + LOT_FILENAME_HISTORICAL.split('\\')[-1],
                                   index=False)


def actual_moskowinvest():
    global driver
    global gkh_df
    global current_date_string
    # global date_to
    # global clf
    # global MID_df
    start_time = time.time()
    print('Обновление данных об актуальных лотах ')
    # считываем данные о прошедших торгах из файла
    print('Попытка получить данные с сайта InvestMoscow')
    # запускаем Chrome браузер в оконном режиме для получения информации с сайта
    # в headless режиме - крашится
#    driver = start_browser_for_parse()
    driver.implicitly_wait(30)
    # url до 26.04.2023 'https://investmoscow.ru/tenders?pageNumber=1&pageSize=100&orderBy=CreateDate&orderAsc=false&objectTypes=7&tenderTypes=13&tenderStatuses=1&tradeForms=45001'
    main_page_url = 'https://investmoscow.ru/tenders?pageNumber=1&pageSize=10&orderBy=RequestEndDate&orderAsc=true&objectTypes=nsi:41:30011568&objectKinds=nsi:tender_type_portal:13&tenderStatus=nsi:tender_status_tender_filter:1&timeToPublicTransportStop.noMatter=true'
    driver.get(main_page_url)
    wait = WebDriverWait(driver, 30)
    try:
        wait.until(
            EC.visibility_of_element_located((By.XPATH,
                                              '//span[@class="uid-text-accent"]')))
        main_page_soup = BeautifulSoup(driver.page_source, 'html.parser')
        qty_lots_str = main_page_soup.find('span', class_="uid-text-accent").text.strip()
    except (Exception, ):
        print('Сайт не отвечает. Попробуйте позже.')
        logging.info('Сайт не отвечает. Попробуйте позже.')
        driver.quit()
        return
    qty_lots = int(re.sub("\D", "", qty_lots_str))
    print('Всего объектов: ', qty_lots)
    logging.info(f'Всего объектов: {qty_lots}')
    if qty_lots == 0:
        return
    url_lst = []
    # считываем данные страниц по актуальным лотам торгам
    for i in trange(int(qty_lots / 100) + 1):
        if len(url_lst) >= qty_lots:
            break
        cur_page_url = 'https://investmoscow.ru/tenders?pageNumber=' + str(i + 1) + '&pageSize=100&' + \
                       'orderBy=RequestEndDate&orderAsc=true&objectTypes=nsi:41:30011568&' + \
                       'objectKinds=nsi:tender_type_portal:13&tenderStatus=nsi:tender_status_tender_filter:1' + \
                        '&timeToPublicTransportStop.noMatter=true'
        driver, url_page_lst = get_url_list_from_page(driver, cur_page_url, is_active=True)
        url_lst = url_lst + url_page_lst
        #url_lst .append(url_page_lst)
    # создаем датафрейм с данными лотов
    lot_df = pd.DataFrame(columns=LOT_FIELDS)
    print("time elapsed: {:.2f}s".format(time.time() - start_time))
    unreadable_urls = []
    error_url_cnt = 0
    driver.implicitly_wait(10)
#---> тут можно добавить ограничение по парсингу первой части
    for url in tqdm(url_lst):
        # считываем страницу с заданным URL, при нескачивании в течении 10 секунд добавляем ссылку в список нескачанных
        try:
            driver, lot_single = parse_lot(driver, url)
            if not (pd.isna(lot_single['start_price'])):
                lot_df = pd.concat([lot_df, lot_single.to_frame().T], ignore_index = True)
            else:
                print(f'Нечитаемая ссылка N{error_url_cnt} {url}')
                print('lot_tag ' + lot_single['lot_tag'])
                logging.info(f'Нечитаемая ссылка N{error_url_cnt} {url}')
                logging.info(f"lot_tag  {lot_single['lot_tag']}")
                unreadable_urls.append(url)
                error_url_cnt += 1
                continue
        except Exception as e:
            print(f'Нечитаемая ссылка N{error_url_cnt} {url}')
            print('lot_tag ' + lot_single['lot_tag'])
            logging.info(f'Нечитаемая ссылка N{error_url_cnt} {url}')
            logging.info(f"lot_tag  {lot_single['lot_tag']}")
            unreadable_urls.append(url)
            error_url_cnt += 1
            continue
    print('Всего новых объектов: ', qty_lots)
    print('Скачалось объектов: ', len(lot_df))
    print('Пропущенных объектов: ', len(unreadable_urls))
    logging.info('Всего новых объектов: ' + str(qty_lots))
    logging.info('Скачалось объектов: ' + str(len(lot_df)))
    logging.info('Пропущенных объектов: ' + str(len(unreadable_urls)))
    # повторно пытаемся скачать ранее нескачанные ссылки с таймаутом 60 секунд
    if len(unreadable_urls) > 0:
        lot_df = reparsing_active_unreadable_urls(driver, lot_df, qty_lots, unreadable_urls)
    # исключаем строки, в прием заявок по которым окончен ранее или сегодня
    today = pd.to_datetime('today')
    lot_df = lot_df[lot_df['finish_application_date'] > today]
#    lot_df.to_csv('time_lot_df.csv',index=False)
#     lot_df = pd.read_csv('time_lot_df.csv', nrows = 70) #, nrows = 90


    # очистка датафрейма от дублирующихся лотов
    if len(lot_df) > 0:
        lot_df = drop_duplicated_lots(lot_df)
    else:
        print('Новые данные отсутвуют')
        logging.info('Новые данные отсутвуют')
    # нормируем и распознаем адреса (если есть новые)
    if len(lot_df) > 0:
        norm_df = recognize_and_normalize_addresses(lot_df)
        # после распознавания адресов - уточняем данные о метро и этажности
        norm_df = is_metro_and_floor_data_complete(norm_df)
    else:
        norm_df = pd.DataFrame()
    norm_df = fill_spaces_in_data(norm_df)
    norm_df[OUTPUT_ACTIVE_COLS].to_csv(LOT_FILENAME, index=False)
# ----> добавляем информацию о торгах с roseltorg.ru
    bidding_df = pd.read_csv(BIDDING_FILENAME)
    bidding_df.id = bidding_df.id.apply(str)
    bidding_df.reestr_num = bidding_df.apply(lambda x: x.reestr_num.strip("'").strip()
                                                        if not x.reestr_num.startswith("'178")
                                                        else str(x.id), axis=1)

    norm_df['reestr_num_lot'] = norm_df.roseltorg_url.apply(lambda x: x.split('/')[-1])
    norm_df = add_bidding_data(norm_df, bidding_df)
    norm_df.drop(columns=['id', 'reestr_num', 'win_price', 'win_name'], inplace=True)
    output_columns = OUTPUT_ACTIVE_COLS + ['participant_qty']
    norm_df[output_columns].to_csv("..\\realty_model_files\\output\\"
                                       + current_date_string[:8] + LOT_FILENAME.split('\\')[-1],
                                   index=False)

def refresh_bidding_df():
    bidding_df = pd.read_csv(BIDDING_FILENAME)
    proc_df = pd.read_excel(r'..\realty_model_files\procedure_folder\procedures.xlsx',
                              # sheet_name='procedures (3)',
                              engine='openpyxl',
                              keep_default_na=False,
                              usecols="A,C,G,S,T,U,V",
                              header=None,
                            skiprows=1,
                            names=['id', 'reestr_num', 'start_price', 'participant_qty',
                                   'win_price', 'applicant_qty', 'win_name'])
    proc_df.applicant_qty = proc_df.applicant_qty.apply(lambda x: int(x) if x != 'Нет'
                                                                            else 0)
    proc_df.participant_qty = proc_df.apply(lambda x: max(x.particpant_qty,
                                                                x.applicant_qty), axis=1)
    proc_df.win_price = proc_df.apply(lambda x: x.start_price
                                                        if (x.win_price == 0 and
                                                            x.participant_qty == 1)
                                                        else x.win_price, axis=1)
    proc_df = proc_df.drop(columns=['applicant_qty', 'start_price'])

    bidding_df = pd.concat([bidding_df, proc_df])
    bidding_df.drop_duplicates(subset=['id'], keep='last', inplace=True)
    bidding_df.to_csv(BIDDING_FILENAME, index=False)


def bidding_def():
    for i in range(18, 47):
        try:
            proc_n_df = pd.read_excel(r'..\realty_model_files\procedure_folder\procedures ('+str(i)+').xlsx',
                                       # sheet_name='procedures (3)',
                                        engine='openpyxl',
                                        keep_default_na=False,
                                        usecols="A,C,G,S,T,U,V",
                                        header=None,
                                        skiprows=1,
                                        names=['id', 'reestr_num', 'start_price', 'participant_qty',
                                               'win_price', 'applicant_qty', 'win_name'])#,
                                        #dtype={'id': int, 'participant_qty': int,
                                         #    'win_price': int, 'applicant_qty': str})
        except:
            print('-', i)
        if i == 18:
            bidding_df = proc_n_df
        else:
            bidding_df = pd.concat([bidding_df, proc_n_df], ignore_index=True)
    bidding_df.applicant_qty = bidding_df.applicant_qty.apply(lambda x: int(x) if x != 'Нет'
                                                                        else 0)
    bidding_df.participant_qty = bidding_df.apply(lambda x: max(x.participant_qty,
                                                                x.applicant_qty), axis=1)
    bidding_df.win_price = bidding_df.apply(lambda x: x.start_price
                                                                if (x.win_price == 0 and
                                                                    x.participant_qty == 1)
                                                                else x.win_price, axis=1)
    bidding_df = bidding_df.drop(columns=['applicant_qty', 'start_price'])
    bidding_df.drop_duplicates(subset=['id'], inplace=True)
    bidding_df.to_csv(r'..\realty_model_files\bidding.csv', index=False)


def get_grid_final_shadow_root(driver):
    shadow_host = WebDriverWait(driver, 5).until(
        EC.visibility_of(driver.find_element(By.TAG_NAME, 'bw-app')))
    shadow_root1 = shadow_host.shadow_root
    root2 = WebDriverWait(driver, 5).until(
        EC.visibility_of(shadow_root1.find_element(By.ID, 'pages')))
    root3 = WebDriverWait(driver, 5).until(
        EC.visibility_of(root2.find_element(By.CLASS_NAME, 'iron-selected')))
    shadow_root3 = root3.shadow_root
    root4 = WebDriverWait(driver, 5).until(
        EC.visibility_of(shadow_root3.find_element(By.ID, 'bwSearchPageTab')))
    shadow_root4 = root4.shadow_root
    root5 = WebDriverWait(driver, 5).until(
        EC.visibility_of(shadow_root4.find_element(By.ID, 'bwSearchPageTabMainContent')))
    shadow_root5 = root5.shadow_root
    root6 = WebDriverWait(driver, 5).until(
        EC.visibility_of(shadow_root5.find_element(By.ID, 'pages')))
    root7 = WebDriverWait(driver, 5).until(
        EC.visibility_of(root6.find_element(By.CSS_SELECTOR, 'bw-panel')))
    root8 = WebDriverWait(driver, 5).until(
        EC.visibility_of(root7.find_element(By.CSS_SELECTOR, 'bw-wide-search-form-grid-page')))
    shadow_root8 = root8.shadow_root
    return driver, shadow_root8

def parse_winner():
    global driver

    driver.quit()
    driver = start_browser_for_parse(pict=True)
    driver.implicitly_wait(3)
    winner_url = 'https://w7.baza-winner.ru/search/9a224416-a516-4e66-afd3-0deb2fb13ea6/list'
    print('')
    driver.get('https://w7.baza-winner.ru/main')#winner_url)
    # ----> нажатие на кнопку "Войти"
    shadow_host = WebDriverWait(driver, 5).until(
        EC.visibility_of(driver.find_element(By.TAG_NAME, 'bw-app')))
    shadow_root1 = shadow_host.shadow_root
    root2 = WebDriverWait(driver, 5).until(
        EC.visibility_of(shadow_root1.find_element(By.ID, 'header')))
    root3 = WebDriverWait(driver, 5).until(
        EC.visibility_of(root2.find_element(By.CSS_SELECTOR, 'bw-user-control')))
    shadow_root3 = root3.shadow_root
    WebDriverWait(driver, 5).until(
        EC.visibility_of(shadow_root3.find_element(By.CSS_SELECTOR, "iron-icon"))).click()
    time.sleep(2)

    act_el  = driver.switch_to.active_element

    act_el.send_keys(WINNER_LOGIN)
    act_el.send_keys(Keys.TAB)
    act_el.send_keys(WINNER_PWD)
    act_el.send_keys(Keys.ENTER)

    time.sleep(10)

    print(driver.current_url)

    winner_df = pd.DataFrame(columns=WINNER_BASE_FIELDS)
    current_position = 0
    prev_position = 0
    total_position = 1
    pbar = tqdm(total=100)
# ---> до парсинга переключаемся на открытую позицию, делаем клик и поднимаемся на начало списка
# (при открытии сохраняется последнее положение на странице)
    time.sleep(1)
    driver, shadow_root8 = get_grid_final_shadow_root(driver)
    root9 = shadow_root8.find_element(By.CLASS_NAME, "ag-center-cols-container")
    row_el = root9.find_element(By.CSS_SELECTOR, "div[role='row']")
    WebDriverWait(driver, 5).until(
        EC.visibility_of(
            row_el.find_element(By.CSS_SELECTOR, "div[aria-colindex='2']"))).click()
    driver.find_element(By.CSS_SELECTOR, 'body').send_keys(Keys.HOME)
    time.sleep(5)
    driver.find_element(By.CSS_SELECTOR, 'body').send_keys(Keys.HOME)
    time.sleep(5)
    driver.find_element(By.CSS_SELECTOR, 'body').send_keys(Keys.HOME)
    time.sleep(5)
    prev_el = row_el
    is_first_iteration = True
# >--------------
    while current_position < total_position + 1:
        try:
            time.sleep(2)
            driver, shadow_root8 = get_grid_final_shadow_root(driver)
            root9 = shadow_root8.find_element(By.CLASS_NAME, "ag-center-cols-container")
            string_els = root9.find_elements(By.CSS_SELECTOR, "div[role='row']")
    #----->поиск номера текущей позиции
            root9_num = shadow_root8.find_element(By.CSS_SELECTOR, "bw-grid-panel-control")
            shadow_root9_num = root9_num.shadow_root
            root10_num = shadow_root9_num.find_element(By.CSS_SELECTOR, "div[id='itemCount']")
            if is_first_iteration:
                position_array = root10_num.text.split('/')
                total_position = int(position_array[1].strip())
                is_first_iteration = False
            pbar.update(round((current_position - prev_position) * 100 / total_position, 2))
            prev_position = current_position
            dict_list = []
            for el in string_els:
                # print(el.get_attribute('aria-rowindex'))
                row = {WINNER_BASE_FIELDS[i]: np.nan for i in range(len(WINNER_BASE_FIELDS))}
                try:
                    row['addr_winner'] = el.find_element(By.CSS_SELECTOR, "div[aria-colindex='9']").text
                except:
                    print('Ошибка опрделения адреса. предыдущая строка')
                    print(prev_el.text)
                try:
                    row['qty_rooms'] = int(el.find_element(By.CSS_SELECTOR, "div[aria-colindex='6']").text)
                except Exception as exc:
                    row['qty_rooms'] = -1
                    print('Ошибка определения комнат ', row['addr_winner'])
                floor_params = el.find_element(By.CSS_SELECTOR, "div[aria-colindex='10']").text
                try:
                    row['addr_floor'] = int(floor_params.split('/')[0])
                    row['total_floors'] = int(floor_params.split('/')[1].split(' ')[0])
                except Exception as exc:
                    row['addr_floor'] = -1
                    row['total_floors'] = -1
                    print('Ошибка определения этажей ', row['addr_winner'], floor_params)
                try:
                    row['obj_square'] = round(float(el.find_element(By.CSS_SELECTOR, "div[aria-colindex='14']")
                                              .text.split('/')[0]), 1)
                except Exception as exc:
                    row['obj_square'] = -1
                    print('Ошибка определения площади ', row['addr_winner'])
                    print(el.find_element(By.CSS_SELECTOR, "div[aria-colindex='14']")
                                              .text.split('/')[0])
                try:
                    row['price'] = int(re.sub(' ', '',
                                        el.find_element(By.CSS_SELECTOR, "div[aria-colindex='16']")
                                              .text.split(',')[0]))
                    row['price_m2'] = round(row['price']/row['obj_square'], 0)
                except Exception as exc:
                    row['price'] = -1
                    row['price_m2'] = -1
                    print('Ошибка определения цены ', row['addr_winner'])
                    print(el.find_element(By.CSS_SELECTOR, "div[aria-colindex='16']")
                          .text)
                try:
                    row['date'] = pd.to_datetime(
                        el.find_element(By.CSS_SELECTOR, "div[aria-colindex='19']").text,
                        infer_datetime_format=True, format='%d.%m.%Y', errors='coerce')
                except Exception as exc:
                    row['date'] = pd.to_datetime('01.01.1970', format='%d.%m.%Y', errors='coerce')

                try:
                    exp_days = el.find_element(By.CSS_SELECTOR, "div[aria-colindex='20']").text
                    if exp_days != '':
                        row['exposition_days'] = int(exp_days)
                    else:
                        row['exposition_days'] = 0
                except Exception as exc:
                    row['exposition_days'] = -1

                row['seller_name'] = el.find_element(By.CSS_SELECTOR, "div[aria-colindex='22']").text
                winner_df = pd.concat([winner_df, pd.DataFrame.from_records([row])], ignore_index=True)
                prev_el = el

            current_position = int(prev_el.get_attribute('aria-rowindex'))
#            print('last_el row_index', current_position)
            driver.find_element(By.CSS_SELECTOR, 'body').send_keys(Keys.PAGE_DOWN)
            driver.find_element(By.CSS_SELECTOR, 'body').send_keys(Keys.PAGE_DOWN)

        except Exception as exc:
            print(traceback.format_exc().split('Stacktrace:')[0])
            print(row)
    pbar.close()
    print(len(winner_df))

# ---> сохраняем дубликаты в файл
    duplicated_winner = winner_df[winner_df.duplicated(subset=['addr_winner', 'qty_rooms', 'addr_floor','total_floors',
                                      'obj_square'], keep=False)]
    duplicated_winner.sort_values(by=['addr_winner', 'qty_rooms', 'addr_floor','total_floors',
                                      'obj_square']).to_csv(r'duplicated_winner.csv', index=False)
# >------------

    winner_df.drop_duplicates(subset=['addr_winner', 'qty_rooms', 'addr_floor','total_floors',
                                      'obj_square'],
                                keep='first', inplace=True)
    print(len(winner_df))
#    driver.quit()
    return winner_df

def winner_def():
    global driver
    global gkh_df
    global new_gkh_df

#   winner_df = parse_winner()
    #cur_winner_df = pd.read_csv(WINNER_FILENAME)
    cur_winner_df = pd.DataFrame()
    winner_df = pd.read_csv(r'test_winner.csv')


    if len(winner_df) > 0:
        winner_addresses = winner_df.drop_duplicates(subset=['addr_winner'])
        gkh_df = pd.read_csv(GKH_BASE_FILENAME)
#        driver = start_browser_for_parse()
        for i, row in winner_addresses[:30].iterrows():
            if row['addr_winner'] not in gkh_df['addr_winner']:
                if not row['addr_winner'].startswith('ЖК'):
                    metro_and_floor_data(row['addr_winner'], None, is_for_winner=True)
        control_new_gkh_df_2gis()
        gkh_df = pd.read_csv(GKH_BASE_FILENAME)
        gkh_df['gkh_address'] = gkh_df.gkh_address.apply(lambda x: str(x).lower())
      #  driver.quit()
        winner_df = winner_df.merge(gkh_df, on='addr_winner', how='left')
        winner_df['addr_norm'] = winner_df['gkh_address']

        winner_df = fill_spaces_in_data(winner_df, is_for_winner=True)
#        output_winner_cols = WINNER_BASE_FIELDS + GKH_FIELDS
        winner_df = pd.concat([cur_winner_df, winner_df], ignore_index=True)
        winner_df = winner_df.drop_duplicates(keep='last').reset_index(drop=True)
        winner_df.to_csv(WINNER_FILENAME, index=False)
    print()

def test_def():
    # lines = ['Москва,Садовническая ул., 77С2',
    #         'Москва, Мира просп., 188Бк3']
    # for line in lines:
    #     ret = metro_and_floor_data(line, None)
    #     #ret = primary_addr_normalize(line)
    #     print(ret)
    df = pd.read_csv(GKH_BASE_FILENAME)
    df['addr_winner'] = np.nan
    df.to_csv(GKH_BASE_FILENAME, index=False)

# Press the green button in the gutter to run the script.
if __name__ == '__main__':

    WINNER_FILENAME = r'..\realty_model_files\winner_data.csv'
    WINNER_LOGIN = "+79800201607"
    WINNER_PWD = "1q2w3e4r5"

    #winner - убрать столбцы addr_street, addr_build_num, добавить addr_norm

    WINNER_BASE_FIELDS = ['addr_winner', 'addr_norm',
                            'qty_rooms', 'addr_floor', 'total_floors', 'obj_square', 'price',
                            'price_m2', 'date', 'is_active', 'seller_name', 'exposition_days']
    LOT_FIELDS = ['lot_tag', 'auct_type', 'object_type', 'cadastr_num', 'addr_string',
                  'addr_street', 'addr_build_num', 'addr_floor_from_title', 'addr_apart_num_from_title',
                  'flat_num', 'qty_rooms', 'addr_floor', 'total_floors', 'obj_square', 'start_price',
                  'start_price_m2','deposit', 'auct_step', 'auct_form', 'start_applications_date',
                  'finish_application_date', 'participant_selection_date',
                  'bidding_date', 'results_date', 'roseltorg_url', 'torgi_url',
                  'inf_sales', 'inf_food_service', 'inf_education',
                  'inf_cult_and_sport', 'inf_consumer_services', 'inf_health_care',
                  'documentation_expl', 'documentation_photo', 'investmoscow_url']
    HISTORICAL_LOT_FIELDS = ['status', 'final_price', 'final_price_m2', 'delta_price']
    LOT_FILENAME = r'..\realty_model_files\realty_model_actual_lot_data.csv'
    LOT_FILENAME_HISTORICAL = r'..\realty_model_files\realty_model_historical_lot_data.csv'
    BIDDING_FILENAME = r'..\realty_model_files\bidding.csv'
    GKH_BASE_FILENAME = r'..\realty_model_files\gkh_base.csv'
    GKH_FIELDS = ['adm_area', 'mun_district', 'gkh_address', 'gkh_total_floors',        'is_total_floors_variable',
                  'geo_lat', 'geo_lon', 'overlap_material', 'skeleton', 'wall_material',
                  'category', 'residents_qty', 'construction_year', 'ceiling_height',
                  'passenger_elevators_qty', 'condition', 'gkh_metro_station',
                  'metro_km', 'metro_min', 'gkh_cadastr_num', 'code_KLADR', 'global_repair_date',
                  'is_renov', 'renov_period',
                  'building_page_url', 'addr_winner']
    OUTPUT_ACTIVE_COLS = ['lot_tag', 'auct_type', 'object_type', 'cadastr_num',
                'adm_area', 'mun_district', 'addr_norm',
                'flat_num', 'qty_rooms', 'addr_floor', 'total_floors', 'is_total_floors_variable',
                'obj_square', 'start_price',
                'start_price_m2','deposit', 'auct_step', 'auct_form', 'start_applications_date',
                'finish_application_date', 'participant_selection_date',
                'bidding_date', 'results_date', 'inf_sales', 'inf_food_service', 'inf_education',
                'inf_cult_and_sport', 'inf_consumer_services', 'inf_health_care',
                'geo_lat', 'geo_lon', 'overlap_material', 'skeleton', 'wall_material',
                'category', 'residents_qty', 'construction_year', 'ceiling_height',
                'passenger_elevators_qty', 'condition', 'gkh_metro_station',
                'metro_km', 'metro_min', 'code_KLADR', 'global_repair_date',
                'is_renov', 'renov_period',
                'documentation_expl', 'documentation_photo',
                'roseltorg_url', 'torgi_url', 'investmoscow_url',
                'building_page_url'
                ]
    SCROLL_PAUSE_TIME = 0.7
    current_date = datetime.datetime.now()
    current_date_string = current_date.strftime('%y_%m_%d_%H_%M')
#    historical_processing()
    print(current_date_string)
    logging.basicConfig(
        level=logging.INFO,
        filename="model_log_" + current_date_string + ".log",
        format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
        datefmt='%H:%M:%S',
    )

    logging.info('Hello')
    gkh_df = pd.DataFrame()
    new_gkh_df = pd.DataFrame()
    date_to = ''
    morph_vocab = MorphVocab()
    addr_extractor = AddrExtractor(morph_vocab)
    choice = 'q'
    min_dict = dict({0: ['&foot_min=10', '<10 мин', '&only_foot=2'],
                     1: ['&foot_min=20', '<20 мин', '&only_foot=2'],
                     2: ['&foot_min=30', '<30 мин', '&only_foot=2'],
                     3: ['', '>30 мин', '']
                     })
    room_dict = dict({0: ['&room1=1', '1'],
                      1: ['&room2=1', '2'],
                      2: ['&room3=1', '3'],
                      3: ['&room4=1&room5=1&room6=1', '4+']
                      })

    driver = start_browser_for_parse()

    try:
        print('Выберите модуль: ')
        print('   1 - Обновить данные по торгам roseltorg')
        print('   2 - актуальные лоты')
        print('   3 - прошедшие лоты')
        print('   4 - WINNER')
        print('   5 - переоформление biddings')
        print('   6 - тест')
        mode_choice = input()
        if mode_choice == '1':
            refresh_bidding_df()
        elif mode_choice == '2':
            actual_moskowinvest()
        elif mode_choice == '3':
            historical_processing()
        elif mode_choice == '4':
            winner_def()
        elif mode_choice == '5':
            bidding_def()
        elif mode_choice == '6':
            test_def()
        driver.quit()
    except Exception as e:
        driver.quit()
        print('Ошибка: ', traceback.format_exc().split('Stacktrace:')[0])
        logging.info(f'Ошибка: {traceback.format_exc().split("Stacktrace:")[0]}')

  #  driver.quit()